#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================
  LİDERFORM + ACCURACE TEK KOŞU EXCEL OLUŞTURUCU  v9.8
  Kullanım : python liderform_tek_kosu_v9_8.py [URL]
  Gereksinim: pip install requests beautifulsoup4 openpyxl
=============================================================
  v9.8 YENİLİKLER (v9.7 üzerine):
  - ELEME ANALİZİ: "ORT. FİNAL SIRA (2K)" sütunu kaldırıldı,
    yerine "STİL" sütunu eklendi (SPRİNTER/GERİDE/ÖN GRUP/OYN.KUR.)
  - ELEME ANALİZİ başlığında BEKLENEN SENARYO (HIZLI/YAVAS) gösterilir
  - Puanlama: %15 STİL/TEMPO uyum puanı → HIZLI tempoda sprinter/geride
    ata +15p, YAVAS tempoda ön grup/oyun kurucuya +15p
  - Yavaş tempoda GERİDE/SPRİNTER karakter + tempo uyumsuz → direkt eleme
  - En kötü Son600 olan 2 at elenir (en iyi koşu bazlı)
  - En iyi Son400 olan 3 at kalın kırmızı yazılır
  - Son koşu verisi boşsa önceki koşu kullanılır (madde 7)
  - DİKKAT VERİ EKSİK uyarısı (v9.7'den korundu)
=============================================================
  - basari_skoru() fonksiyonu: tuple karşılaştırma ile
    daha sağlam en iyi koşu seçimi (final_sira, -s400_sn)
  - Cinsiyet tespiti: parse_ana_sayfa() kısrak/dişi/aygır
    tespiti + at sözlüğüne "cinsiyet" alanı eklendi
  - Irk tespiti: race_info tipinden "ARAP" / "İNGİLİZ" ayrımı
  - Arap atları eleme: en iyi koşudaki mesafe farkı >200m
    ise direkt elenir
  - İngiliz atları eleme: karışık cinsiyet koşusunda dişi
    at direkt elenir
  - Eleme bazı değiştirildi: v9.6 "son" koşu → v9.7 "en_iyi"
    koşu (basari_skoru ile seçilen)
  - Karma koşu atlamak daha sağlam is_karma() ile
  - veri_eksik uyarısı + kilo/KGS kontrolleri korundu
  - Tüm TEMPO & SENARYO, sınıf avantajı, ACCURACE sayfaları
    korundu
=============================================================
  VERİ sayfası   → A2'den başlar
  ANALİZ sayfası → Bilgi A1:B5, veriler F6'dan başlar
  ACCURACE sayfası → Her at için son+önceki koşu hız verisi
  ELEME ANALİZİ → en_iyi koşu bazlı puanlama (ırk+cinsiyet)
  TEMPO & SENARYO → Oyun kurucu / sprinter analizi
=============================================================
"""

import re, sys, os, time, requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Renkler ─────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
GREEN      = "1E6B2E"
ALT_ROW    = "EBF3FB"
ALT_GREEN  = "E8F5E9"
WHITE      = "FFFFFF"
HEADER_ACC = "1B5E20"   # Accurace sayfası başlık

def _border(c="BDD7EE"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def style_hdr(cell, bg=MID_BLUE):
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border("FFFFFF")

def style_dat(cell, even=False, bold=False, left=False, bg=None):
    if bg is None:
        bg = ALT_ROW if even else WHITE
    cell.font      = Font(name="Arial", size=10, bold=bold)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left" if left else "center", vertical="center")
    cell.border    = _border()

# ─── HTTP ─────────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;"
        "q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8"
    ),
    "Accept-Language":  "tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding":  "gzip, deflate, br",
    "Connection":       "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest":   "document",
    "Sec-Fetch-Mode":   "navigate",
    "Sec-Fetch-Site":   "none",
    "Sec-Fetch-User":   "?1",
    "Cache-Control":    "max-age=0",
    "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
    "sec-ch-ua-mobile":   "?0",
    "sec-ch-ua-platform": '"Windows"',
    "Referer": "https://www.liderform.com.tr/",
    "DNT": "1",
}

# ── Playwright ile HTML çek (bot engelini aşar) ────────────
def _fetch_html_playwright(url):
    """Gerçek Chromium tarayıcısıyla sayfayı açar — Cloudflare/bot engeline karşı."""
    from playwright.sync_api import sync_playwright
    import os, base64

    chromium_path = os.environ.get("CHROMIUM_PATH", None)

    html = None
    with sync_playwright() as p:
        launch_opts = dict(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--single-process",
                "--window-size=1920,1080",
                "--disable-web-security",
                "--disable-features=IsolateOrigins,site-per-process",
            ]
        )
        if chromium_path and os.path.exists(chromium_path):
            launch_opts["executable_path"] = chromium_path

        browser = p.chromium.launch(**launch_opts)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="tr-TR",
            timezone_id="Europe/Istanbul",
            viewport={"width": 1920, "height": 1080},
            extra_http_headers={
                "Accept-Language": "tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
                "Referer": "https://www.liderform.com.tr/",
                "DNT": "1",
            }
        )
        # navigator.webdriver'ı ve diğer bot izlerini gizle
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['tr-TR','tr','en-US','en']});
            window.chrome = {runtime: {}};
            Object.defineProperty(navigator, 'permissions', {
                get: () => ({ query: () => Promise.resolve({ state: 'granted' }) })
            });
        """)
        page = ctx.new_page()

        # Önce ana sayfayı ziyaret et (cookie ve oturum al)
        try:
            print("         Ana sayfa ziyaret ediliyor (cookie için)...")
            page.goto("https://www.liderform.com.tr/", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(2000)
            # Mouse hareketi simülasyonu (bot değiliz diyoruz)
            page.mouse.move(500, 300)
            page.wait_for_timeout(500)
        except Exception as e:
            print(f"         Ana sayfa hatası (devam): {e}")

        # Asıl sayfayı aç
        print(f"         Hedef sayfa açılıyor: {url[:60]}")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(3000)

        # Sayfa başlığını logla (bot koruması tespiti için)
        title = page.title()
        print(f"         Sayfa başlığı: {title}")

        # Screenshot al ve logla (debug için)
        try:
            screenshot = page.screenshot(type="png")
            b64 = base64.b64encode(screenshot).decode()
            print(f"         SCREENSHOT_B64_START:{b64[:100]}...SCREENSHOT_B64_END")
        except Exception as e:
            print(f"         Screenshot alınamadı: {e}")

        html = page.content()
        browser.close()
    return html

# ── requests fallback (local/VPS ortamı için) ─────────────
_SESSION = requests.Session()
_SESSION.headers.update(HEADERS)

def _fetch_html_requests(url, deneme=3, bekleme=5):
    global _SESSION
    for i in range(1, deneme + 1):
        try:
            if i == 1:
                try:
                    _SESSION.get("https://www.liderform.com.tr/", timeout=15)
                    time.sleep(1)
                except Exception:
                    pass
            r = _SESSION.get(url, timeout=60, allow_redirects=True)
            if r.status_code == 403:
                _SESSION = requests.Session()
                _SESSION.headers.update(HEADERS)
                time.sleep(bekleme)
                continue
            r.raise_for_status()
            return r.text
        except Exception as e:
            if i < deneme:
                time.sleep(bekleme)
            else:
                raise
    raise Exception(f"Sayfa {deneme} denemede yüklenemedi: {url}")

# ── Ana fetch fonksiyonu — önce playwright dene ───────────
def fetch_soup(url, deneme=3, bekleme=5):
    print(f"         Bağlanıyor: {url[:70]}")
    # Playwright mevcut mu?
    try:
        import playwright
        _playwright_ok = True
    except ImportError:
        _playwright_ok = False

    html = None
    if _playwright_ok:
        try:
            print("         Playwright ile açılıyor...")
            html = _fetch_html_playwright(url)
            print("         Playwright başarılı ✓")
        except Exception as e:
            print(f"         Playwright hatası: {e} — requests'e geçiliyor...")
            html = None

    if html is None:
        print("         requests ile deneniyor...")
        html = _fetch_html_requests(url, deneme=deneme, bekleme=bekleme)

    if not html:
        raise Exception("Sayfa içeriği boş geldi.")

    soup = BeautifulSoup(html, "html.parser")

    # 403 / bot koruması kontrolü
    title = soup.title.string if soup.title else ""
    if "403" in title or "Access Denied" in title or "Forbidden" in title:
        raise Exception(
            "Liderform.com.tr erişim engelledi (403). "
            "Lütfen birkaç dakika sonra tekrar deneyin."
        )
    if "Just a moment" in (soup.get_text()[:200] or ""):
        raise Exception(
            "Cloudflare doğrulaması aktif. Lütfen 1-2 dakika bekleyip tekrar deneyin."
        )

    return soup

# ─── URL Meta ─────────────────────────────────────────────
AY = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
      "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

def url_meta(url):
    m = re.search(r"/program/(\d{4}-\d{2}-\d{2})/([^/]+)/(\d+)", url)
    if not m: return {}
    y, mo, d = m.group(1).split("-")
    return {
        "tarih_iso": m.group(1),
        "tarih_tr":  f"{int(d)} {AY[int(mo)]} {y}",
        "sehir":     m.group(2).replace("-"," ").title(),
        "sehir_raw": m.group(2),
        "kosu_no":   m.group(3),
    }

# ─── Koşu başlık bilgisi ──────────────────────────────────
def race_info(soup):
    info = {"tip":"", "pist":"", "mesafe":"", "kategori":""}
    h2 = soup.find("h2")
    if not h2: return info
    txt = h2.get_text(" ", strip=True)
    for pat, key in [
        (r"([A-ZÇĞİÖŞÜa-zçğışöü]+-\d+(?:/D)?)", "tip"),
        (r"(Sentetik|Çim|Kum)",                   "pist"),
        (r"(\d{3,4})\s*m",                        "mesafe"),
        (r"(\d+\s+Yaşl[ıi]\s+\S+)",               "kategori"),
    ]:
        mm = re.search(pat, txt, re.I)
        if mm: info[key] = mm.group(1).strip()
    if info["mesafe"]: info["mesafe"] += " m"
    return info

# ─── LİDERFORM PROGRAM SAYFASI PARSE (cinsiyet eklendi) ──
def parse_ana_sayfa(soup):
    atlar = []
    at_divleri = [
        d for d in soup.find_all("div")
        if "font-RobotoRegular" in str(d.get("class", ""))
        and "text-[#181D27]"    in str(d.get("class", ""))
    ]
    print(f"       Bulunan at satırı: {len(at_divleri)}")

    for div in at_divleri:
        metin = div.get_text(" ", strip=True)

        a_tag = div.find("a", href=re.compile(r"/istatistik/at/\d+"))
        if not a_tag: continue
        at_adi = a_tag.get_text(strip=True)
        if not at_adi or len(at_adi) < 2 or at_adi.startswith("("): continue

        # CİNSİYET tespiti — kısrak/dişi = dişi; aygır/iğdiş/erkek = erkek
        cinsiyet = ""
        lower_metin = metin.lower()
        if any(k in lower_metin for k in ("kısrak", "dişi")):
            cinsiyet = "dişi"
        elif re.search(r"\bk\b", lower_metin):
            cinsiyet = "dişi"
        elif any(k in lower_metin for k in ("aygır", "erkek", "iğdiş")):
            cinsiyet = "erkek"
        elif re.search(r"\b[ei]\b", lower_metin):
            cinsiyet = "erkek"

        # NO
        at_no = 0
        btn = div.find("button")
        if btn:
            t = btn.get_text(strip=True)
            if re.match(r"^\d{1,2}$", t): at_no = int(t)
        if at_no == 0:
            m = re.match(r"^(\d{1,2})\s", metin)
            if m: at_no = int(m.group(1))

        # AT ADI'ndan sonraki metin
        pos         = metin.find(at_adi)
        after       = metin[pos + len(at_adi):].strip() if pos != -1 else metin
        after_clean = re.sub(r"Önceki\s*:\s*\d+", "", after).strip()

        # GP & HP
        gp, hp = "", ""
        sayilar = re.findall(r"\b(\d{1,3})\b", after_clean[:60])
        if len(sayilar) >= 1: gp = int(sayilar[0])
        if len(sayilar) >= 2: hp = int(sayilar[1])

        # KG
        kg = ""
        taki_m = re.search(
            r"\b((?:(?:DB|GKR|DS)\s+)*(?:[A-Z]+\s+)*KG|KG\s+SK|KG|-)\b",
            after_clean, re.I
        )
        if taki_m:
            sonrasi = after_clean[taki_m.end():taki_m.end()+25]
            km = re.search(r"\b(\d{2,3}(?:\.\d+)?)\b", sonrasi)
            if km:
                v = float(km.group(1))
                if 50 <= v <= 70:
                    kg = int(v) if v == int(v) else v
        if kg == "":
            bulunan = []
            for num in re.findall(r"\b(\d{2,3}(?:\.\d+)?)\b", after_clean[:120]):
                v = float(num)
                if 50 <= v <= 70:
                    bulunan.append(int(v) if v == int(v) else v)
            # GP ve HP zaten parse edildi, bunları sırayla çıkar
            kalan = list(bulunan)
            for cikar in [gp, hp]:
                if cikar in kalan:
                    kalan.remove(cikar)
            if kalan:
                kg = kalan[0]
            elif bulunan:
                # GP/HP ile örtüşse bile son çare olarak al
                kg = bulunan[-1]

        # JOKEY
        jokey = ""
        j_tag = div.find("a", href=re.compile(r"/istatistik/jokey/"))
        if j_tag:
            raw   = j_tag.get_text(strip=True)
            jokey = re.sub(r"\s*\d+\s*/\s*\d+.*$|\s*AP$", "", raw).strip()

        # YRC - boşluklu ve bitişik format
        yrc = ""
        ym = re.search(r"\b(\d)\s+[A-ZÇĞİÖŞÜa-zçğışöü]\s+[A-ZÇĞİÖŞÜa-zçğışöü]\b", after_clean)
        if ym:
            yrc = int(ym.group(1))
        else:
            ym2 = re.search(r"\b(\d)([a-zA-ZÇĞİÖŞÜçğışöü]{2})\b", after_clean)
            if ym2: yrc = int(ym2.group(1))

        # ST
        st = ""
        yrc_pos = re.search(r"\b\d\s+[A-ZÇĞİÖŞÜa-zçğışöü]\s+[A-ZÇĞİÖŞÜa-zçğışöü]\b", after_clean)
        if not yrc_pos:
            yrc_pos = re.search(r"\b\d[a-zA-ZÇĞİÖŞÜçğışöü]{2}\b", after_clean)
        if yrc_pos:
            sonrasi = after_clean[yrc_pos.end():]
            st_nums = re.findall(r"\b(\d{1,2})\b", sonrasi[:15])
            if st_nums:
                v = int(st_nums[0])
                if 1 <= v <= 25: st = v

        # EN İYİ DERECE ve PİST ŞEHİR + ACCURACE LİNK BİLGİSİ
        en_iyi     = ""
        pist_sehir = ""
        acc_tarih  = ""
        acc_sehir  = ""
        acc_kosu   = ""

        PIST_MAP  = {"Sen":"Sentetik","Çim":"Çim","Kum":"Kum","No":"Normal","Ne":"Nemli"}
        SEHIR_MAP = {
            "İst":"İstanbul","Bur":"Bursa","Ada":"Adana",
            "Ank":"Ankara","İzm":"İzmir","Ela":"Elazığ",
            "Diy":"Diyarbakır","Koc":"Kocaeli"
        }
        # Accurace şehir kodu eşleşmeleri (büyük harf)
        SEHIR_ACC = {
            "İst":"ISTANBUL","Bur":"BURSA","Ada":"ADANA",
            "Ank":"ANKARA","İzm":"IZMIR","Ela":"ELAZIG",
            "Diy":"DIYARBAKIR","Koc":"KOCAELI"
        }

        for lnk in div.find_all("a", href=re.compile(r"/sonuclar/")):
            txt  = lnk.get_text(strip=True)
            href = lnk.get("href", "")
            # "54/2.06.70/Sen/İst" formatı
            dm = re.match(r"([\d.]+)/([\d.:]+)/([^/]+)/([^/]+)$", txt)
            if dm:
                en_iyi     = dm.group(2)
                pist_txt   = dm.group(3)
                sehir_txt  = dm.group(4)
                pist_long  = PIST_MAP.get(pist_txt, pist_txt)
                sehir_long = SEHIR_MAP.get(sehir_txt, sehir_txt)
                pist_sehir = f"{pist_long} / {sehir_long}"

                # /sonuclar/2026-03-28/istanbul/4 → accurace bilgileri
                hm = re.search(r"/sonuclar/(\d{4}-\d{2}-\d{2})/([^/]+)/(\d+)", href)
                if hm:
                    acc_tarih = hm.group(1)                          # 2026-03-28
                    acc_sehir = SEHIR_ACC.get(sehir_txt,
                                sehir_txt.upper().replace("İ","I"))  # ISTANBUL
                    acc_kosu  = hm.group(3)                          # 4
                break

        # KGS
        kgs = ""
        km = re.search(r"\b(\d{1,3})\s+%\s*[\d.]+\s*\(\d+\)", metin)
        if km: kgs = int(km.group(1))

        # AGF %
        agf_pct = ""
        am = re.search(r"%\s*([\d.]+)\s*\(\d+\)", metin)
        if am: agf_pct = "%" + am.group(1)

        atlar.append({
            "no": at_no, "adi": at_adi, "cinsiyet": cinsiyet,
            "gp": gp, "hp": hp, "kg": kg,
            "jokey": jokey, "yrc": yrc, "st": st,
            "en_iyi_derece": en_iyi, "pist_sehir": pist_sehir,
            "kgs": kgs, "agf_pct": agf_pct,
            # Accurace linki için
            "acc_tarih": acc_tarih,
            "acc_sehir": acc_sehir,
            "acc_kosu":  acc_kosu,
        })

    goren, temiz = set(), []
    for at in sorted(atlar, key=lambda x: x["no"]):
        if at["adi"] not in goren and at["no"] > 0:
            goren.add(at["adi"])
            temiz.append(at)
    return temiz

# ─── PERFORMANS SAYFASINDAN SON + ÖNCEKİ KOŞU LİNKİ ─────
def son_kosu_linklerini_cek(tarih, sehir, kosu_no):
    """
    /program/performans/TARIH/SEHIR/KOSU sayfasından
    her atın son 2 geçerli koşu linkini çeker.
    KARMA (KAR) koşular atlanır, bir önceki alınır.
    Son 2 koşudan daha başarılı olanı (ilk 4'e girmiş veya daha iyi son400) seçer.

    Döner: {at_adi: {
        "son":     {"tarih","sehir","kosu","mesafe"},   <- daha başarılı koşu
        "onceki":  {"tarih","sehir","kosu","mesafe"},   <- diğer koşu
        "secilen_indeks": 0 veya 1  (hangi sıranın seçildiği)
    }}
    """
    url = f"https://liderform.com.tr/program/performans/{tarih}/{sehir}/{kosu_no}"
    try:
        soup = fetch_soup(url, deneme=2, bekleme=3)
    except Exception as e:
        print(f"         UYARI: Performans sayfası alınamadı: {e}")
        return {}

    SEHIR_ACC = {
        "İst":"ISTANBUL","Bur":"BURSA","Ada":"ADANA",
        "Ank":"ANKARA","İzm":"IZMIR","Ela":"ELAZIG",
        "Diy":"DIYARBAKIR","Koc":"KOCAELI",
        "istanbul":"ISTANBUL","bursa":"BURSA","adana":"ADANA",
        "ankara":"ANKARA","izmir":"IZMIR",
    }

    def satir_karma_mi(row):
        """Satırda 'KAR' veya 'KARMA' geçiyor mu? (hem td hem link metni)"""
        txt = row.get_text(" ", strip=True).upper()
        if "KARMA" in txt or re.search(r"\bKAR\b", txt):
            return True
        for a in row.find_all("a"):
            atxt = a.get_text(strip=True).upper()
            if "KARMA" in atxt or re.search(r"\bKAR\b", atxt):
                return True
        return False

    def satir_parse(row):
        """Bir tablo satırından acc bilgilerini çıkar."""
        sonuc_link = None
        for a in row.find_all("a", href=True):
            if "/sonuclar/" in a["href"]:
                sonuc_link = a["href"]
                break
        if not sonuc_link: return {}

        hm = re.search(r"/sonuclar/(\d{4}-\d{2}-\d{2})/([^/]+)/(\d+)", sonuc_link)
        if not hm: return {}

        acc_tarih = hm.group(1)
        sehir_raw = hm.group(2)
        acc_kosu  = hm.group(3)

        # Sehir kısaltmasını bul (td[1] = "İzm", "İst" vb.)
        tds = row.find_all("td")
        sehir_kisa = tds[1].get_text(strip=True) if len(tds) > 1 else ""
        acc_sehir  = SEHIR_ACC.get(sehir_kisa,
                     SEHIR_ACC.get(sehir_raw,
                     sehir_raw.upper().replace("İ","I").replace("Ş","S")
                     .replace("Ğ","G").replace("Ü","U").replace("Ö","O")
                     .replace("Ç","C")))

        # Mesafe: td[3] = "1900K", "1300S"
        mesafe_str = ""
        if len(tds) >= 4:
            m = re.match(r"(\d+)", tds[3].get_text(strip=True))
            if m: mesafe_str = m.group(1) + "m"

        # Derece / final sırası: tüm td'lerde küçük sayıyı bul (1-20 arası = sıra)
        final_sira = None
        for td in tds[4:]:  # İlk 4 td'yi atla (tarih, şehir, koşu, mesafe)
            txt = td.get_text(strip=True)
            sm  = re.match(r"^(\d{1,2})$", txt)
            if sm:
                v = int(sm.group(1))
                if 1 <= v <= 20:
                    final_sira = v
                    break

        # KG: tüm td'lerde 50-70 arasındaki sayıyı bul
        son_kg = ""
        for td in tds:
            txt = td.get_text(strip=True)
            km  = re.match(r"^(\d{2,3}(?:\.\d)?)$", txt)
            if km:
                v = float(km.group(1))
                if 50 <= v <= 70:
                    son_kg = int(v) if v == int(v) else v
                    break

        return {
            "tarih":      acc_tarih,
            "sehir":      acc_sehir,
            "kosu":       acc_kosu,
            "mesafe":     mesafe_str,
            "kg":         son_kg,
            "final_sira": final_sira,  # koşudaki bitiş sırası
        }

    # At linklerini sıralı bul
    at_links = []
    seen = set()
    for tag in soup.find_all("a", href=re.compile(r"/istatistik/at/\d+")):
        metin = tag.get_text(strip=True)
        if not metin or metin.startswith("(") or len(metin) < 2: continue
        aid = re.search(r"/at/(\d+)", tag["href"]).group(1)
        if aid in seen: continue
        seen.add(aid)
        at_links.append(tag)

    sonuc = {}
    for at_tag in at_links:
        at_adi = at_tag.get_text(strip=True)
        tablo  = at_tag.find_next("table")
        if not tablo: continue
        rows = tablo.find_all("tr")
        if len(rows) < 2: continue

        # KARMA olmayan ilk 2 koşuyu topla
        gecerli_kosular = []
        for row in rows[1:]:  # rows[0] başlık
            if satir_karma_mi(row):
                print(f"           {at_adi}: KARMA koşu atlandı → bir önceki alınıyor")
                continue
            bilgi = satir_parse(row)
            if bilgi:
                gecerli_kosular.append(bilgi)
            if len(gecerli_kosular) >= 2:
                break

        if not gecerli_kosular:
            continue

        bilgi_1 = gecerli_kosular[0]   # En son geçerli koşu
        bilgi_2 = gecerli_kosular[1] if len(gecerli_kosular) >= 2 else {}

        # Daha başarılı koşuyu seç:
        # 1. İlk 4'e girmiş olanı tercih et
        # 2. İkisi de ilk 4'teyse veya ikisi de dışındaysa: final sıra daha iyi olanı
        def basari_skoru(b):
            """Düşük skor = daha başarılı."""
            if not b: return 999
            fs = b.get("final_sira")
            if fs is None: return 500
            return fs

        sk1 = basari_skoru(bilgi_1)
        sk2 = basari_skoru(bilgi_2)

        # ilk 4 kontrolü
        ilk4_1 = sk1 <= 4
        ilk4_2 = sk2 <= 4 if bilgi_2 else False

        if ilk4_1 and not ilk4_2:
            # 1. koşu daha başarılı
            secilen = bilgi_1
            diger   = bilgi_2
            secilen_lbl = "SON (ilk4)"
        elif ilk4_2 and not ilk4_1:
            # 2. koşu daha başarılı
            secilen = bilgi_2
            diger   = bilgi_1
            secilen_lbl = "ÖNCEKİ (ilk4)"
        elif sk1 <= sk2:
            # İkisi de ilk4 veya ikisi de dışında; sırası daha iyi olanı al
            secilen = bilgi_1
            diger   = bilgi_2
            secilen_lbl = f"SON (sıra:{sk1})"
        else:
            secilen = bilgi_2
            diger   = bilgi_1
            secilen_lbl = f"ÖNCEKİ (sıra:{sk2})"

        print(f"         {at_adi}:")
        print(f"           Seçilen: {secilen.get('tarih','?')}/{secilen.get('sehir','?')}/{secilen.get('kosu','?')} ({secilen.get('mesafe','?')}) [{secilen_lbl}]")
        if diger:
            print(f"           Diğer  : {diger.get('tarih','?')}/{diger.get('sehir','?')}/{diger.get('kosu','?')} ({diger.get('mesafe','?')})")

        sonuc[at_adi] = {
            "son":    secilen,   # daha başarılı koşu (analiz için kullanılacak)
            "onceki": diger,     # diğer koşu
        }

    return sonuc

# ─── ACCURACE VERİSİ ÇEK ─────────────────────────────────
def parse_accurace_soup(soup, at_adi_hedef=None):
    """
    Accurace sayfasını parse eder.
    at_adi_hedef verilirse sadece o atın verisini döner.
    Döner: (at_verileri_dict, mesafeler_list, mesafe_metre_str)
      mesafe_metre_str: "2000m", "1600m" gibi son mesafe
    """
    tablo = soup.find("table")
    if not tablo:
        return {}, [], ""

    rows = tablo.find_all("tr")
    if not rows:
        return {}, [], ""

    # Başlık satırından mesafeleri al
    header_cells = rows[0].find_all(["td","th"])
    mesafeler = []
    for cell in header_cells[1:]:
        txt = cell.get_text(strip=True)
        m = re.match(r"(\d+m)\.", txt)
        if m: mesafeler.append(m.group(1))

    # Son mesafe = koşu mesafesi
    son_mesafe = mesafeler[-1] if mesafeler else ""

    # At satırları
    at_verileri = {}
    for row in rows[1:]:
        cells = row.find_all(["td","th"])
        if not cells: continue
        ilk  = cells[0].get_text(strip=True)
        no_m = re.match(r"^(\d{1,2})", ilk)
        if not no_m: continue
        at_adi = ilk[len(no_m.group(0)):].strip()

        sure_dict = {}
        for i, cell in enumerate(cells[1:]):
            if i >= len(mesafeler): break
            txt = cell.get_text(strip=True)
            sm  = re.match(r"(.+?)\[(\d+)\]", txt)
            if sm:
                sure_dict[mesafeler[i]] = {
                    "sure": sm.group(1),
                    "sira": int(sm.group(2))
                }
        at_verileri[at_adi] = {"sureler": sure_dict}

    return at_verileri, mesafeler, son_mesafe

def fetch_accurace_per_horse(horses, son_kosu_map):
    """
    Her at için son koşu VE bir önceki koşunun Accurace verisini çeker.
    son_kosu_map: {at_adi: {"son":{...}, "onceki":{...}}}
    Döner: {at_adi: {
        "son":    {"sureler","mesafeler","son_mesafe","acc_url","son400","son600"},
        "onceki": {"sureler","mesafeler","son_mesafe","acc_url","son400","son600"}
    }}
    """
    url_cache = {}
    sonuc     = {}

    def acc_veri_cek(bilgi, at_adi, etiket):
        """Bir koşu için Accurace verisi çek, son400/son600 hesapla."""
        if not bilgi:
            return {"sureler":{},"mesafeler":[],"son_mesafe":"","acc_url":"",
                    "son400":"","son600":""}

        acc_tarih = bilgi.get("tarih","")
        acc_sehir = bilgi.get("sehir","")
        acc_kosu  = bilgi.get("kosu","")
        acc_mes   = bilgi.get("mesafe","")

        if not acc_tarih or not acc_sehir or not acc_kosu:
            print(f"           {at_adi} [{etiket}]: Bilgi yok, atlanıyor.")
            return {"sureler":{},"mesafeler":[],"son_mesafe":"","acc_url":"",
                    "son400":"","son600":""}

        url = f"https://accurace.net/network/{acc_tarih}/{acc_sehir}/{acc_kosu}/summary"

        if url not in url_cache:
            print(f"           {at_adi} [{etiket}]: {url}")
            try:
                soup = fetch_soup(url, deneme=2, bekleme=3)
                at_verileri, mesafeler, son_mesafe = parse_accurace_soup(soup)
                url_cache[url] = (at_verileri, mesafeler, son_mesafe)
            except Exception as e:
                print(f"           UYARI: {e}")
                url_cache[url] = ({}, [], "")

        at_verileri, mesafeler, son_mesafe = url_cache[url]
        if acc_mes: son_mesafe = acc_mes

        # Bu atın verisini bul (fuzzy)
        at_verisi = at_verileri.get(at_adi, None)
        if at_verisi is None:
            for k in at_verileri:
                if k.strip().upper() == at_adi.strip().upper():
                    at_verisi = at_verileri[k]
                    break
        if at_verisi is None:
            at_verisi = {"sureler": {}}

        sureler = at_verisi.get("sureler", {})
        s400, s600 = hesapla_son_400_600(sureler, son_mesafe)

        time.sleep(0.2)
        return {
            "sureler":    sureler,
            "mesafeler":  mesafeler,
            "son_mesafe": son_mesafe,
            "acc_url":    url,
            "son400":     s400,
            "son600":     s600,
        }

    for horse in horses:
        at_adi = horse["adi"]
        bilgiler = son_kosu_map.get(at_adi, {})

        son_veri    = acc_veri_cek(bilgiler.get("son",{}),    at_adi, "SON")
        onceki_veri = acc_veri_cek(bilgiler.get("onceki",{}), at_adi, "ÖNCEKİ")

        # ── basari_skoru ile en iyi koşuyu seç ──────────────────
        # Tuple karşılaştırma: (final_sira, -s400_sn)
        # final_sira küçük = iyi, s400_sn küçük (hızlı) = iyi → negatif alınır
        def basari_skoru(veri):
            mes = veri.get("son_mesafe", "")
            final_sira = veri.get("sureler", {}).get(mes, {}).get("sira", 99) if mes else 99
            s400_sn = sure_to_sec(veri.get("son400", ""))
            return (final_sira, -s400_sn if s400_sn else 0)

        skor_son    = basari_skoru(son_veri)
        skor_onceki = basari_skoru(onceki_veri)

        # Madde 7: Eğer son koşu dereceleri boşsa önceki koşuyu kullan
        son_veri_bos = (not son_veri.get("sureler") and not son_veri.get("son400") and not son_veri.get("son600"))
        if son_veri_bos and onceki_veri.get("sureler"):
            en_iyi_veri = onceki_veri
            en_iyi_etik = "ÖNCEKİ (son boş)"
            print(f"           {at_adi}: Son koşu verisi boş → ÖNCEKİ KOŞU kullanılıyor")
        elif skor_son <= skor_onceki:
            en_iyi_veri = son_veri
            en_iyi_etik = "SON"
            print(f"           {at_adi}: En iyi → SON KOŞU (sıra:{skor_son[0]}, s400:{-skor_son[1]:.2f}sn)")
        else:
            en_iyi_veri = onceki_veri
            en_iyi_etik = "ÖNCEKİ"
            print(f"           {at_adi}: En iyi → ÖNCEKİ KOŞU (sıra:{skor_onceki[0]}, s400:{-skor_onceki[1]:.2f}sn)")
        # Form karşılaştırması: Son 400 ve Son 600 farkı (küçük = hızlı = YÜKSELİŞ)
        form_400 = ""
        form_600 = ""
        fark_400_str = ""
        fark_600_str = ""

        s400_son    = sure_to_sec(son_veri.get("son400",""))
        s400_onceki = sure_to_sec(onceki_veri.get("son400",""))
        s600_son    = sure_to_sec(son_veri.get("son600",""))
        s600_onceki = sure_to_sec(onceki_veri.get("son600",""))

        if s400_son is not None and s400_onceki is not None:
            fark = round(s400_son - s400_onceki, 2)
            fark_400_str = f"{fark:+.2f}sn"
            form_400 = "YÜKSELİŞ" if fark < 0 else ("STABIL" if fark == 0 else "DÜŞÜŞ")

        if s600_son is not None and s600_onceki is not None:
            fark = round(s600_son - s600_onceki, 2)
            fark_600_str = f"{fark:+.2f}sn"
            form_600 = "YÜKSELİŞ" if fark < 0 else ("STABIL" if fark == 0 else "DÜŞÜŞ")

        # Genel form değerlendirmesi (ikisi de varsa birlikte değerlendir)
        if form_400 and form_600:
            if form_400 == "YÜKSELİŞ" and form_600 == "YÜKSELİŞ":
                genel_form = "⬆ YÜKSELİŞ"
            elif form_400 == "DÜŞÜŞ" and form_600 == "DÜŞÜŞ":
                genel_form = "⬇ DÜŞÜŞ"
            elif "YÜKSELİŞ" in (form_400, form_600):
                genel_form = "↗ KISMI YÜKSELİŞ"
            elif "DÜŞÜŞ" in (form_400, form_600):
                genel_form = "↘ KISMI DÜŞÜŞ"
            else:
                genel_form = "→ STABIL"
        elif form_400:
            genel_form = ("⬆ YÜKSELİŞ" if form_400=="YÜKSELİŞ"
                          else ("⬇ DÜŞÜŞ" if form_400=="DÜŞÜŞ" else "→ STABIL"))
        elif form_600:
            genel_form = ("⬆ YÜKSELİŞ" if form_600=="YÜKSELİŞ"
                          else ("⬇ DÜŞÜŞ" if form_600=="DÜŞÜŞ" else "→ STABIL"))
        else:
            genel_form = ""

        sonuc[at_adi] = {
            "son":         son_veri,
            "onceki":      onceki_veri,
            "en_iyi":      en_iyi_veri,
            "en_iyi_etik": en_iyi_etik,
            "form_400":    form_400,
            "form_600":    form_600,
            "fark_400":    fark_400_str,
            "fark_600":    fark_600_str,
            "genel_form":  genel_form,
        }

    return sonuc

# ─── EXCEL ────────────────────────────────────────────────
# VERİ / ANALİZ sayfası sütunları
COLS = [
    ("NO",             5),
    ("AT ADI",        18),
    ("GP",             7),
    ("HP",             7),
    ("KG",             6),
    ("JOKEY",         14),
    ("YRC",            6),
    ("ST",             6),
    ("EN İYİ DERECE", 16),
    ("PİST ŞEHİR",    18),
    ("KGS",            7),
    ("AGF %",         10),
]
HDR = [c[0] for c in COLS]
WID = [c[1] for c in COLS]

def horse_vals(h):
    return [
        h.get("no",""),            h.get("adi",""),
        h.get("gp",""),            h.get("hp",""),
        h.get("kg",""),            h.get("jokey",""),
        h.get("yrc",""),           h.get("st",""),
        h.get("en_iyi_derece",""), h.get("pist_sehir",""),
        h.get("kgs",""),           h.get("agf_pct",""),
    ]

def write_block(ws, meta, rinfo, horses, start_row, start_col):
    nc = len(HDR)
    ec = start_col + nc - 1
    title = (
        f"{meta.get('sehir','').upper()}  {meta.get('kosu_no','')}. KOŞU  |  "
        f"{meta.get('tarih_tr','')}  |  "
        f"{rinfo.get('tip','')}  |  {rinfo.get('pist','')}  |  "
        f"{rinfo.get('mesafe','')}  |  {rinfo.get('kategori','')}"
    )
    tc = ws.cell(row=start_row, column=start_col, value=title)
    tc.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    tc.fill      = PatternFill("solid", start_color=DARK_BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row,   end_column=ec)
    ws.row_dimensions[start_row].height = 22

    hr = start_row + 1
    for ci, h in enumerate(HDR, start_col):
        style_hdr(ws.cell(row=hr, column=ci, value=h))
    ws.row_dimensions[hr].height = 18

    for ri, horse in enumerate(horses):
        rn, even = hr + 1 + ri, ri % 2 == 0
        for ci, val in enumerate(horse_vals(horse), start_col):
            cell = ws.cell(row=rn, column=ci, value=val)
            style_dat(cell, even=even, bold=(ci==start_col+1), left=(ci==start_col+1))
        ws.row_dimensions[rn].height = 16

    for ci, w in enumerate(WID, start_col):
        ws.column_dimensions[get_column_letter(ci)].width = w

def write_accurace_sheet(ws, meta, rinfo, horses, acc_data, mesafeler):
    """
    ACCURACE sayfası:
    Her at için bir blok:
      - At adı başlığı (koyu yeşil)
      - Mesafe | Süre | Sıra
    """
    if not mesafeler:
        ws["A1"] = "Accurace verisi bulunamadı."
        return

    # Sayfa başlığı
    title = (
        f"ACCURACE HIZLANMA VERİSİ  |  "
        f"{meta.get('sehir','').upper()} {meta.get('kosu_no','')}. KOŞU  |  "
        f"{meta.get('tarih_tr','')}  |  {rinfo.get('mesafe','')} {rinfo.get('pist','')}"
    )
    nc = len(mesafeler) * 2 + 1   # AT ADI + (SÜRE+SIRA) * mesafe sayısı
    tc = ws.cell(row=1, column=1, value=title)
    tc.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    tc.fill      = PatternFill("solid", start_color=HEADER_ACC)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(nc,8))
    ws.row_dimensions[1].height = 24

    # Sütun başlıkları
    ws.cell(row=2, column=1, value="AT").apply(lambda c: style_hdr(c, GREEN))
    col = 2
    for mes in mesafeler:
        c1 = ws.cell(row=2, column=col,   value=f"{mes} SÜRE")
        c2 = ws.cell(row=2, column=col+1, value=f"{mes} SIRA")
        style_hdr(c1, GREEN)
        style_hdr(c2, GREEN)
        col += 2
    ws.row_dimensions[2].height = 18

    # Her at için veri satırı (NO sırasına göre)
    for ri, horse in enumerate(horses):
        at_adi = horse["adi"]
        rn     = 3 + ri
        even   = ri % 2 == 0

        # AT ADI
        c = ws.cell(row=rn, column=1, value=f"{horse['no']}-{at_adi}")
        style_dat(c, even=even, bold=True, left=True)
        ws.column_dimensions["A"].width = 20

        # Accurace verisi
        acc = acc_data.get(at_adi, {})
        sureler = acc.get("sureler", {}) if acc else {}

        col = 2
        for mes in mesafeler:
            veri  = sureler.get(mes, {})
            sure  = veri.get("sure", "")
            sira  = veri.get("sira", "")

            c1 = ws.cell(row=rn, column=col,   value=sure)
            c2 = ws.cell(row=rn, column=col+1, value=sira)
            style_dat(c1, even=even)
            style_dat(c2, even=even)

            # 1. sıra ise yeşil arka plan
            if sira == 1:
                c1.fill = PatternFill("solid", start_color="C8E6C9")
                c2.fill = PatternFill("solid", start_color="C8E6C9")
                c1.font = Font(name="Arial", size=10, bold=True, color="1B5E20")
                c2.font = Font(name="Arial", size=10, bold=True, color="1B5E20")

            ws.column_dimensions[get_column_letter(col)].width   = 12
            ws.column_dimensions[get_column_letter(col+1)].width = 8
            col += 2

        ws.row_dimensions[rn].height = 16

# Hücreye stil uygulama yardımcısı (lambda yerine)
def _apply_style(cell, style_fn, **kwargs):
    style_fn(cell, **kwargs)
    return cell

def sure_to_sec(s):
    """
    Süre string -> saniye (float).
    "27''33"    -> 27.33
    "1'25''00"  -> 85.00
    "2'15''90"  -> 135.90
    """
    if not s: return None
    s = str(s).strip()
    # Format: (?:DK')? TAM '' KES
    # Örn: "1'25''00"  veya  "27''33"
    m = re.match(r"^(?:(\d+)')?(\d+)''(\d+)$", s)
    if m:
        dk  = int(m.group(1)) if m.group(1) else 0
        return dk * 60 + int(m.group(2)) + int(m.group(3)) / 100
    # Fallback: düz float
    try:
        return float(s.replace(",", "."))
    except:
        return None

def sec_to_str(sec):
    """
    Saniye -> süre string.
    27.36  -> "27''36"
    85.00  -> "1'25''00"
    Format: [DK']TAM''KES  (sure_to_sec ile tam uyumlu)
    """
    if sec is None: return ""
    sec = round(sec, 2)
    if sec >= 60:
        dk  = int(sec // 60)
        san = sec - dk * 60
        tam = int(san)
        kes = round((san - tam) * 100)
        return f"{dk}'{tam:02d}''{kes:02d}"
    else:
        tam = int(sec)
        kes = round((sec - tam) * 100)
        return f"{tam:02d}''{kes:02d}"

def interpolate_sure(hedef_m, sureler_dict):
    """
    Lineer interpolasyon ile hedef mesafeye karşılık gelen süreyi tahmin eder.
    sureler_dict: {"400m": {"sure":"27''33","sira":1}, ...}
    hedef_m: int (metre cinsinden)
    Döner: float (saniye) veya None
    """
    # Mevcut mesafe -> saniye tablosu
    mesafe_sec = {}
    for key, veri in sureler_dict.items():
        m = re.match(r"^(\d+)m$", key)
        if m and veri.get("sure"):
            sec = sure_to_sec(veri["sure"])
            if sec is not None:
                mesafe_sec[int(m.group(1))] = sec

    if not mesafe_sec: return None

    mesafeler = sorted(mesafe_sec.keys())

    # Tam eşleşme → doğrudan döndür
    if hedef_m in mesafe_sec:
        return mesafe_sec[hedef_m]

    # Lineer interpolasyon için komşuları bul
    altlar = [m for m in mesafeler if m < hedef_m]
    ustler = [m for m in mesafeler if m > hedef_m]

    if not altlar or not ustler:
        return None  # Ekstrapolasyon yapmıyoruz

    m1, m2 = max(altlar), min(ustler)
    t1, t2 = mesafe_sec[m1], mesafe_sec[m2]

    oran = (hedef_m - m1) / (m2 - m1)
    return t1 + oran * (t2 - t1)

def hesapla_son_400_600(sureler, son_mesafe):
    """
    Son 400m ve son 600m sürelerini hesaplar.
    Formül:
        Son 400 = T(D) - T(D-400)
        Son 600 = T(D) - T(D-600)
    D-400 veya D-600 veride yoksa lineer interpolasyon kullanır.

    Örnekler:
        1200m: Son600 = T(1200)-T(600)  → doğrudan
        1300m: Son600 = T(1300)-T(700)  → interpolasyon (600m-800m arası)
        1900m: Son600 = T(1900)-T(1300) → interpolasyon (1200m-1600m arası)
    """
    if not son_mesafe: return "", ""
    son_m = int(re.sub(r"[^0-9]","", son_mesafe))

    # Final süre
    final_sec = interpolate_sure(son_m, sureler)
    if final_sec is None: return "", ""

    # Son 400
    son_400 = ""
    ref_400_sec = interpolate_sure(son_m - 400, sureler)
    if ref_400_sec is not None:
        son_400 = sec_to_str(round(final_sec - ref_400_sec, 2))

    # Son 600
    son_600 = ""
    ref_600_sec = interpolate_sure(son_m - 600, sureler)
    if ref_600_sec is not None:
        son_600 = sec_to_str(round(final_sec - ref_600_sec, 2))

    return son_400, son_600

def build_excel(meta, rinfo, horses, acc_data, mesafeler, fname, son_kosu_map=None, web_mode=False):
    if son_kosu_map is None: son_kosu_map = {}
    _web_sonuclar   = []   # ELEME sonuçları (web_mode için)
    _web_tempo      = {}   # TEMPO kararı (web_mode için)
    wb = Workbook()

    # ── VERİ sayfası ──
    ws_v = wb.active
    ws_v.title = "VERİ"
    write_block(ws_v, meta, rinfo, horses, start_row=2, start_col=1)

    # ── ANALİZ sayfası ──
    ws_a = wb.create_sheet("ANALİZ")
    for i, (lbl, val) in enumerate([
        ("Tarih :",  meta.get("tarih_tr","")),
        ("Şehir :",  meta.get("sehir","")),
        ("Koşu :",   meta.get("kosu_no","") + ". Koşu"),
        ("Pist :",   rinfo.get("pist","") + " – " + rinfo.get("mesafe","")),
        ("Tip :",    rinfo.get("tip","") + " | " + rinfo.get("kategori","")),
    ], 1):
        lc = ws_a.cell(row=i, column=1, value=lbl)
        lc.font = Font(name="Arial", bold=True, size=10, color=MID_BLUE)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc = ws_a.cell(row=i, column=2, value=val)
        vc.font = Font(name="Arial", size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws_a.row_dimensions[i].height = 16
    ws_a.column_dimensions["A"].width = 10
    ws_a.column_dimensions["B"].width = 22
    write_block(ws_a, meta, rinfo, horses, start_row=6, start_col=6)

    # ── ACCURACE sayfası ──────────────────────────────────────
    ws_acc = wb.create_sheet("ACCURACE")

    ORANGE       = "E65100"
    PURPLE       = "4A148C"
    YUKSELIS_CLR = "1B5E20"
    DUSUS_CLR    = "B71C1C"
    STABIL_CLR   = "1565C0"
    SON_BG       = "E3F2FD"
    ONCEKI_BG    = "FFF9C4"
    FORM_BG      = "F3E5F5"
    SEHIR_GOSTER = {
        "ISTANBUL":"İstanbul","BURSA":"Bursa","ADANA":"Adana",
        "ANKARA":"Ankara","IZMIR":"İzmir","ELAZIG":"Elazığ",
        "DIYARBAKIR":"Diyarbakır","KOCAELI":"Kocaeli"
    }

    # Tüm mesafeleri topla (son + önceki)
    tum_mesafeler = []
    for veri in acc_data.values():
        for kt in ["son","onceki"]:
            for mes in veri.get(kt,{}).get("mesafeler",[]):
                if mes not in tum_mesafeler:
                    tum_mesafeler.append(mes)
    tum_mesafeler.sort(key=lambda x: int(re.sub(r"[^0-9]","",x)))

    FIXED = [("NO-AT ADI",22),("KOŞU",9),("MESAFE",9),("ŞEHİR",12)]
    total_cols = len(FIXED) + len(tum_mesafeler)*2 + 5

    # Satır 1: Büyük başlık
    acc_title = (
        f"ACCURACE SON KOŞU HIZLANMA VERİSİ  |  "
        f"{meta.get('sehir','').upper()} {meta.get('kosu_no','')}. KOŞU  |  "
        f"{meta.get('tarih_tr','')}"
    )
    tc = ws_acc.cell(row=1, column=1, value=acc_title)
    tc.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    tc.fill      = PatternFill("solid", start_color=HEADER_ACC)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_acc.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=max(total_cols, 10))
    ws_acc.row_dimensions[1].height = 24

    # Satır 2: Sütun başlıkları
    for ci, (lbl, w) in enumerate(FIXED, 1):
        style_hdr(ws_acc.cell(row=2, column=ci, value=lbl), bg=GREEN)
        ws_acc.column_dimensions[get_column_letter(ci)].width = w

    col = len(FIXED) + 1
    for mes in tum_mesafeler:
        style_hdr(ws_acc.cell(row=2, column=col,   value=f"{mes}\nSÜRE"), bg=GREEN)
        style_hdr(ws_acc.cell(row=2, column=col+1, value=f"{mes}\nSIRA"), bg=GREEN)
        ws_acc.column_dimensions[get_column_letter(col)].width   = 10
        ws_acc.column_dimensions[get_column_letter(col+1)].width = 6
        col += 2

    col_son400=col; col_son600=col+1; col_fark400=col+2; col_fark600=col+3; col_form=col+4

    for c, lbl, bg in [
        (col_son400,"SON\n400m",ORANGE),(col_son600,"SON\n600m",ORANGE),
        (col_fark400,"FARK\n400m",PURPLE),(col_fark600,"FARK\n600m",PURPLE),
        (col_form,"FORM",DARK_BLUE)
    ]:
        style_hdr(ws_acc.cell(row=2, column=c, value=lbl), bg=bg)
        ws_acc.column_dimensions[get_column_letter(c)].width = 13 if c==col_form else 11
    ws_acc.row_dimensions[2].height = 28

    # Satır 3+: Her at için 2 satır (SON / ÖNCEKİ)
    def dolu_satir(rn, at_label, etiket, veri, bg_clr, at_adi_gercek, kosu_tipi_key, acc=None):
        def hucre(r, c, val, bold=False, color="000000", bg=None):
            cx = ws_acc.cell(row=r, column=c, value=val)
            cx.font = Font(name="Arial", size=10, bold=bold, color=color)
            cx.fill = PatternFill("solid", start_color=bg or bg_clr)
            cx.alignment = Alignment(horizontal="center" if c>1 else "left",
                                     vertical="center")
            cx.border = _border()

        hucre(rn, 1, at_label, bold=True)
        hucre(rn, 2, etiket, bold=(etiket=="SON KOŞU"))
        cx_etiket = ws_acc.cell(row=rn, column=2)
        cx_etiket.font      = Font(name="Arial", size=8, bold=(etiket=="SON KOŞU"))
        cx_etiket.alignment = Alignment(horizontal="left", vertical="center")
        hucre(rn, 3, veri.get("son_mesafe",""))

        # ŞEHİR — son_kosu_map["at_adi"]["son"/"onceki"]["sehir"]
        bilgi_map = son_kosu_map.get(at_adi_gercek, {})
        sehir_acc = bilgi_map.get(kosu_tipi_key, {}).get("sehir","")
        hucre(rn, 4, SEHIR_GOSTER.get(sehir_acc, sehir_acc.title() if sehir_acc else ""))

        # Mesafe süre/sıra
        sureler = veri.get("sureler", {})
        c = len(FIXED)+1
        for mes in tum_mesafeler:
            v    = sureler.get(mes, {})
            sure = v.get("sure","")
            sira = v.get("sira","")
            c1 = ws_acc.cell(row=rn, column=c,   value=sure)
            c2 = ws_acc.cell(row=rn, column=c+1, value=sira if sira!="" else "")
            for cx in [c1,c2]:
                cx.fill = PatternFill("solid", start_color=bg_clr)
                cx.font = Font(name="Arial", size=10)
                cx.alignment = Alignment(horizontal="center", vertical="center")
                cx.border = _border()
            if sira == 1:
                for cx in [c1,c2]:
                    cx.fill = PatternFill("solid", start_color="C8E6C9")
                    cx.font = Font(name="Arial",size=10,bold=True,color=YUKSELIS_CLR)
            elif mes == "800m" and isinstance(sira, int) and sira < 4:
                c2.font = Font(name="Arial", size=10, bold=True, color="B71C1C")
                c2.fill = PatternFill("solid", start_color="FFCDD2")
            c += 2

        # Son400, Son600
        s400 = veri.get("son400","")
        s600 = veri.get("son600","")
        hucre(rn, col_son400, s400, bold=bool(s400), color=ORANGE if s400 else "000000")
        hucre(rn, col_son600, s600, bold=bool(s600), color=ORANGE if s600 else "000000")

        # FARK ve FORM → sadece SON KOŞU satırına yaz
        if etiket == "SON KOŞU" and acc:
            def fark_renk(form):
                return YUKSELIS_CLR if form=="YÜKSELİŞ" else (DUSUS_CLR if form=="DÜŞÜŞ" else STABIL_CLR)

            fark_400 = acc.get("fark_400","")
            form_400 = acc.get("form_400","")
            if fark_400:
                cx = ws_acc.cell(row=rn, column=col_fark400, value=fark_400)
                cx.font = Font(name="Arial",size=10,bold=True,color=fark_renk(form_400))
                cx.fill = PatternFill("solid", start_color=bg_clr)
                cx.alignment = Alignment(horizontal="center",vertical="center")
                cx.border = _border()
            else:
                hucre(rn, col_fark400, "")

            fark_600 = acc.get("fark_600","")
            form_600 = acc.get("form_600","")
            if fark_600:
                cx = ws_acc.cell(row=rn, column=col_fark600, value=fark_600)
                cx.font = Font(name="Arial",size=10,bold=True,color=fark_renk(form_600))
                cx.fill = PatternFill("solid", start_color=bg_clr)
                cx.alignment = Alignment(horizontal="center",vertical="center")
                cx.border = _border()
            else:
                hucre(rn, col_fark600, "")

            genel = acc.get("genel_form","")
            if genel:
                if "YÜKSELİŞ" in genel: fclr,fbg = YUKSELIS_CLR,"C8E6C9"
                elif "DÜŞÜŞ"   in genel: fclr,fbg = DUSUS_CLR,  "FFCDD2"
                else:                    fclr,fbg = STABIL_CLR,  "E3F2FD"
                cx = ws_acc.cell(row=rn, column=col_form, value=genel)
                cx.font = Font(name="Arial",size=10,bold=True,color=fclr)
                cx.fill = PatternFill("solid", start_color=fbg)
                cx.alignment = Alignment(horizontal="center",vertical="center")
                cx.border = _border()
            else:
                hucre(rn, col_form, "")
        else:
            # Önceki koşu satırında FARK/FORM boş
            for ci in [col_fark400, col_fark600, col_form]:
                hucre(rn, ci, "")

        ws_acc.row_dimensions[rn].height = 16

    rn = 3
    for horse in horses:
        at_adi = horse["adi"]
        acc    = acc_data.get(at_adi, {})
        if not acc:
            rn += 2
            continue
        at_label = f"{horse['no']}-{at_adi}"
        dolu_satir(rn,   at_label, "SON KOŞU",    acc.get("son",{}),    SON_BG,    at_adi, "son",    acc)
        dolu_satir(rn+1, "",       "ÖNCEKİ KOŞU", acc.get("onceki",{}), ONCEKI_BG, at_adi, "onceki", None)
        # Atlar arasına kalın çizgi (önceki koşu satırının altına)
        thick_side = Side(style="medium", color="000000")
        for ci in range(1, total_cols + 1):
            cell = ws_acc.cell(row=rn+1, column=ci)
            old_border = cell.border
            cell.border = Border(
                left=old_border.left,
                right=old_border.right,
                top=old_border.top,
                bottom=thick_side
            )
        rn += 2

    if not tum_mesafeler:
        ws_acc.cell(row=3, column=1, value="Accurace verisi bulunamadı.")

    # Koşu mesafesi — hem TEMPO ÖN HESABI hem ELEME ANALİZİ kullanır
    try:
        kosu_mesafe = int(re.sub(r"[^0-9]","", rinfo.get("mesafe","0")))
    except:
        kosu_mesafe = 0

    # ══════════════════════════════════════════════════════════
    # ── TEMPO UYUM ÖN HESABI (ELEME'ye girdi sağlar) ──────────
    # ══════════════════════════════════════════════════════════
    def _get_sure_sn_pre(veri, mes):
        s = veri.get("sureler", {}).get(mes, {}).get("sure", "")
        return sure_to_sec(s) if s else None

    # Mesafeye göre tempo ölçüm noktası: 1600m+ → 1000m, kısa → 800m
    _tempo_mes   = "1000m" if kosu_mesafe >= 1600 else "800m"
    _TEMPO_TOL   = 1.5 if _tempo_mes == "1000m" else 1.0   # 1000m için tolerans biraz geniş
    _BASARI_SIRA = 4

    _onde_sureler = []
    for horse in horses:
        acc_ = acc_data.get(horse["adi"], {})
        son_    = acc_.get("son", {})
        onceki_ = acc_.get("onceki", {})
        for veri_ in [son_, onceki_]:
            sira_ = veri_.get("sureler", {}).get(_tempo_mes, {}).get("sira", "")
            sure_ = _get_sure_sn_pre(veri_, _tempo_mes)
            if isinstance(sira_, int) and sira_ <= 3 and sure_:
                _onde_sureler.append(sure_)

    _tahmini_tempo_sn = sum(_onde_sureler) / len(_onde_sureler) if _onde_sureler else None

    # Her at için tempo uyum skoru ön hesabı
    _tempo_uyum = {}   # at_adi → {"uyum_skoru": int, "uyumlu": bool, "uyumsuz": bool}
    for horse in horses:
        at_adi_ = horse["adi"]
        acc_    = acc_data.get(at_adi_, {})
        son_    = acc_.get("son", {})
        onceki_ = acc_.get("onceki", {})

        def _final_sira(veri):
            mes = veri.get("son_mesafe","") or veri.get("mesafe","")
            return veri.get("sureler", {}).get(mes, {}).get("sira","")

        uyum_skoru_ = 0
        uyumsuz_kez = 0
        for veri_, sf_ in [(son_, _final_sira(son_)), (onceki_, _final_sira(onceki_))]:
            su_ = _get_sure_sn_pre(veri_, _tempo_mes)
            if su_ is None or _tahmini_tempo_sn is None:
                continue
            kaldirabilir = su_ <= _tahmini_tempo_sn + _TEMPO_TOL
            basarili_    = isinstance(sf_, int) and sf_ <= _BASARI_SIRA
            if kaldirabilir:
                uyum_skoru_ += 2 if basarili_ else 1
            else:
                uyumsuz_kez += 1

        _tempo_uyum[at_adi_] = {
            "uyum_skoru": uyum_skoru_,
            "uyumlu":  uyum_skoru_ >= 2,
            "uyumsuz": uyumsuz_kez >= 2,
        }

    # ══════════════════════════════════════════════════════════
    # ── TEMPO KARARI HESABI (ELEME'den önce, tam versiyon) ────
    # ══════════════════════════════════════════════════════════
    # Bu hesap TEMPO & SENARYO sayfasıyla birebir aynı mantığı kullanır.
    # Sonuç olan tempo_karar ve _at_roller ELEME ANALİZİ sayfasında da kullanılır.

    _TEMPO_MES_PRE = "1000m" if kosu_mesafe >= 1600 else "800m"
    _TEMPO_TOL_PRE = 1.5 if _TEMPO_MES_PRE == "1000m" else 1.0
    _BASARI_SIRA_PRE = 4

    _at_tempo_pre = []
    for horse in horses:
        _at_adi = horse["adi"]
        _acc = acc_data.get(_at_adi, {})
        _son = _acc.get("son", {})
        _onc = _acc.get("onceki", {})

        def _gs(_v, _m): return _v.get("sureler",{}).get(_m,{}).get("sira","")
        def _gsn(_v, _m):
            _s = _v.get("sureler",{}).get(_m,{}).get("sure","")
            return sure_to_sec(_s) if _s else None

        _s8s = _gs(_son, _TEMPO_MES_PRE)
        _s8o = _gs(_onc, _TEMPO_MES_PRE)
        _sur8s = _gsn(_son, _TEMPO_MES_PRE)
        _sur8o = _gsn(_onc, _TEMPO_MES_PRE)

        _sf_son = _son.get("sureler",{}).get(_son.get("son_mesafe",""),{}).get("sira","")
        _sf_onc = _onc.get("sureler",{}).get(_onc.get("son_mesafe",""),{}).get("sira","")

        _s400_str = _acc.get("son",{}).get("son400","") or ""
        _s400_sn  = sure_to_sec(_s400_str) if _s400_str else None
        _s400_str2 = _acc.get("onceki",{}).get("son400","") or ""

        _oyun_k  = (_s8s == 1 and _s8o == 1)
        _pot_k   = (_s8s == 1 or _s8o == 1)
        _sure_baz = (not _pot_k and isinstance(_s8s, int) and _s8s <= 3 and _sur8s is not None)
        _on_grup = (isinstance(_s8s, int) and _s8s <= 4)
        _geride  = (isinstance(_s8s, int) and _s8s >= 5)

        _best_sure = min([s for s in [_sur8s, _sur8o] if s is not None], default=None)

        _at_tempo_pre.append({
            "adi": _at_adi, "no": horse["no"],
            "s8_son": _s8s, "s8_onc": _s8o,
            "sf_son": _sf_son, "sf_onc": _sf_onc,
            "sure8_son": _sur8s, "sure8_onc": _sur8o,
            "best_sure": _best_sure,
            "son400_sn": _s400_sn,
            "oyun_kurucu": _oyun_k,
            "pot_oyun_kurucu": _pot_k,
            "sure_bazli_aday": _sure_baz,
            "on_grup": _on_grup,
            "geride": _geride,
            "sprinter": False,
        })

    # Tahmini tempo
    _onde_srs = []
    for _t in _at_tempo_pre:
        for _su, _si in [(_t["sure8_son"], _t["s8_son"]), (_t["sure8_onc"], _t["s8_onc"])]:
            if isinstance(_si, int) and _si <= 3 and _su is not None:
                _onde_srs.append(_su)
        if _t["sure_bazli_aday"] and _t["sure8_son"] is not None:
            _onde_srs.append(_t["sure8_son"])
    _tahmini_tempo_sn_pre = sum(_onde_srs)/len(_onde_srs) if _onde_srs else None

    # Sprinter tespiti
    _s400_listesi_pre = [_t["son400_sn"] for _t in _at_tempo_pre if _t["son400_sn"]]
    _ort_s400_pre = sum(_s400_listesi_pre)/len(_s400_listesi_pre) if _s400_listesi_pre else None
    for _t in _at_tempo_pre:
        _t["sprinter"] = _t["geride"] and bool(
            _t["son400_sn"] and _ort_s400_pre and _t["son400_sn"] < _ort_s400_pre
        )

    # sure_bazli_aday yoksa en hızlısını pot. kurucu yap
    _oyun_k_list = [_t for _t in _at_tempo_pre if _t["oyun_kurucu"]]
    _pot_k_list  = [_t for _t in _at_tempo_pre if _t["pot_oyun_kurucu"] and not _t["oyun_kurucu"]]
    _sure_baz_list = sorted(
        [_t for _t in _at_tempo_pre if _t["sure_bazli_aday"]],
        key=lambda x: x["sure8_son"] if x["sure8_son"] else 999
    )
    if not _oyun_k_list and not _pot_k_list and _sure_baz_list:
        for _t in _at_tempo_pre:
            if _t["adi"] == _sure_baz_list[0]["adi"]:
                _t["pot_oyun_kurucu"] = True
                _t["sure_bazli"] = True
        _pot_k_list = [_sure_baz_list[0]]

    # Tempo kararı (TEMPO & SENARYO ile birebir aynı mantık)
    if len(_oyun_k_list) >= 2:
        _tempo_karar = "HIZLI"
    elif len(_oyun_k_list) == 1:
        _rakipler = [_t for _t in _pot_k_list
                     if isinstance(_t["s8_son"], int) and _t["s8_son"] <= 3]
        _tempo_karar = "HIZLI" if _rakipler else "YAVAS"
    elif len(_pot_k_list) >= 2:
        _tempo_karar = "HIZLI"
    elif len(_pot_k_list) == 1:
        _tempo_karar = "YAVAS"
    else:
        _tempo_karar = "BELIRSIZ"

    # _at_roller: ELEME ANALİZİ stil tespiti için
    _at_roller = {_t["adi"]: _t for _t in _at_tempo_pre}

    # ══════════════════════════════════════════════════════════
    # ── SINIF AVANTAJI HESABI ──────────────────────────────────
    # ══════════════════════════════════════════════════════════
    bugun_hp_listesi = sorted(
        [h.get("hp", 0) for h in horses if isinstance(h.get("hp"), (int, float)) and h.get("hp", 0) > 0],
        reverse=True
    )[:5]
    bugun_ort_hp = sum(bugun_hp_listesi) / len(bugun_hp_listesi) if bugun_hp_listesi else 0

    _sinif_avantaj = {}
    for horse in horses:
        at_adi_ = horse["adi"]
        at_hp   = horse.get("hp", 0) or 0
        try:
            at_hp = float(at_hp)
        except:
            at_hp = 0

        if bugun_ort_hp > 0 and at_hp > 0:
            fark_hp = at_hp - bugun_ort_hp
            if fark_hp >= 5:
                sinif_lbl  = f"SINIF AVANTAJI (HP:{at_hp:.0f}, ort:{bugun_ort_hp:.0f}, +{fark_hp:.0f})"
                sinif_var  = True
                sinif_seviye = "yüksek" if fark_hp >= 10 else "orta"
            elif fark_hp <= -5:
                sinif_lbl  = f"SINIF DEZAVANTAJI (HP:{at_hp:.0f}, ort:{bugun_ort_hp:.0f}, {fark_hp:.0f})"
                sinif_var  = False
                sinif_seviye = "dezavantaj"
            else:
                sinif_lbl  = f"Eşit sınıf (HP:{at_hp:.0f}, ort:{bugun_ort_hp:.0f})"
                sinif_var  = False
                sinif_seviye = "esit"
        else:
            sinif_lbl    = "HP verisi yok"
            sinif_var    = False
            sinif_seviye = "bilinmiyor"

        _sinif_avantaj[at_adi_] = {
            "var":     sinif_var,
            "lbl":     sinif_lbl,
            "seviye":  sinif_seviye,
            "at_hp":   at_hp,
            "fark_hp": at_hp - bugun_ort_hp if bugun_ort_hp > 0 else 0,
        }

    # ══════════════════════════════════════════════════════════
    # ── ELEME ANALİZİ sayfası ─────────────────────────────────
    # ══════════════════════════════════════════════════════════
    ws_e = wb.create_sheet("ELEME ANALİZİ")

    # ── Ağırlıklar (toplam 100 puan) ──────────────────────────
    W_SON400 = 30   # Bitiş hızı — son 400m (birincil)
    W_SON600 = 20   # Bitiş hızı — son 600m
    W_FORM   = 20   # Form trendi
    W_MESAFE = 15   # Mesafe uyumu
    W_STIL   = 15   # Koşu stili — tempo bazlı (HIZLI→sprinter/geride, YAVAS→ön grup/numara)
    W_800    = W_STIL  # geriye dönük uyumluluk için alias

    # ── 1. Pass: ham verileri topla (en_iyi koşu bazlı) ──────
    ham = []
    for horse in horses:
        at_adi  = horse["adi"]
        acc     = acc_data.get(at_adi, {})
        en_iyi  = acc.get("en_iyi", {})  # basari_skoru ile seçilmiş en iyi koşu
        s400_str = en_iyi.get("son400","")
        s600_str = en_iyi.get("son600","")
        ham.append({
            "horse": horse, "acc": acc, "en_iyi": en_iyi,
            "s400_str": s400_str, "s400_sn": sure_to_sec(s400_str) if s400_str else None,
            "s600_str": s600_str, "s600_sn": sure_to_sec(s600_str) if s600_str else None,
        })

    # ── Ortalamalar ───────────────────────────────────────────
    g400 = [t["s400_sn"] for t in ham if t["s400_sn"]]
    g600 = [t["s600_sn"] for t in ham if t["s600_sn"]]
    ort_400 = sum(g400)/len(g400) if g400 else None
    ort_600 = sum(g600)/len(g600) if g600 else None

    # Mesafe eşiği: 1600m+ koşularda kilo kritik
    KG_ESIK       = 62.0   # bu üzeri "aşırı kilo" sayılır
    KGS_ESIK      = 100    # 100 gün ve üzeri start almamış → direkt elenir
    KG_FARK_ESIK  = 5.0    # ortalamadan 5+ kg fazla = aşırı kilo
    bugun_kg_listesi = [
        float(h.get("kg", 0) or 0)
        for h in horses
        if h.get("kg", 0) and str(h.get("kg","")).replace(".","").isdigit()
    ]
    ort_kg = sum(bugun_kg_listesi) / len(bugun_kg_listesi) if bugun_kg_listesi else 0

    # ── IRK TESPİTİ ───────────────────────────────────────────
    # race_info tipinden: "ARAP-4500" → is_arap=True; "Gİ-4000" → is_ingiliz=True
    kosu_tipi  = rinfo.get("tip", "")
    is_arap    = bool(re.search(r"\bARAP\b", kosu_tipi, re.I))
    is_ingiliz = bool(re.search(r"\b(İNGİLİZ|INGILIZ|Gİ|GI)\b", kosu_tipi, re.I))

    # ── CİNSİYET KARIŞIMI KONTROLÜ (sadece İngiliz) ───────────
    erkek_var = any(h.get("cinsiyet") == "erkek" for h in horses)
    disi_var  = any(h.get("cinsiyet") == "dişi"  for h in horses)
    karma_cinsiyet = erkek_var and disi_var

    if karma_cinsiyet and is_ingiliz:
        disi_sayisi  = sum(1 for h in horses if h.get("cinsiyet") == "dişi")
        erkek_sayisi = sum(1 for h in horses if h.get("cinsiyet") == "erkek")
        print(f"       ℹ İngiliz koşusu karma cinsiyet: {disi_sayisi} dişi, {erkek_sayisi} erkek → dişiler elenecek")

    # ── 2. Pass: puanlama ─────────────────────────────────────
    sonuclar = []
    for t in ham:
        horse  = t["horse"]
        acc    = t["acc"]
        en_iyi = t["en_iyi"]
        at_adi = horse["adi"]
        puan   = 0
        detaylar = []
        eleme_nedenleri = []

        # — KGS kontrolü: 100+ gün start almamış → direkt elenir —
        kgs = horse.get("kgs", 0) or 0
        try:
            kgs = int(kgs)
        except:
            kgs = 0
        if kgs >= KGS_ESIK:
            eleme_nedenleri.append(f"Uzun süre start almamış ({kgs} gün)")
            detaylar.append(f"❌ {kgs} gündür start yok → form bilinmiyor")

        # — Kilo kontrolü: 1600m+ koşularda aşırı kilo → direkt elenir —
        at_kg = 0
        kg_artis = 0
        try:
            at_kg = float(str(horse.get("kg", 0) or 0).replace(",", "."))
        except:
            at_kg = 0
        if kosu_mesafe >= 1600 and at_kg > 0:
            # Son koşudaki kg'ı al
            son_kosu_bilgi = son_kosu_map.get(at_adi, {}).get("son", {})
            son_kg = son_kosu_bilgi.get("kg", 0) or 0
            try:
                son_kg = float(son_kg)
            except:
                son_kg = 0
            kg_artis = (at_kg - son_kg) if son_kg > 0 else 0

            if at_kg > 60 and kg_artis >= 2:
                eleme_nedenleri.append(f"Aşırı kilo + artış ({at_kg}kg, önceki:{son_kg}kg, +{kg_artis:.1f}kg artış) → 1600m+ dezavantaj")
                detaylar.append(f"❌ Kilo: {at_kg}kg (önceki:{son_kg}kg, +{kg_artis:.1f}kg artış, 1600m+ için kritik)")
            elif at_kg > 60:
                detaylar.append(f"⚠️ Kilo: {at_kg}kg (60kg üzeri, artış yok veya bilinmiyor)")
            elif at_kg > 58:
                detaylar.append(f"⚠️ Kilo: {at_kg}kg (1600m+ için dikkat)")


        # — Derece / veri eksikliği uyarısı —
        veri_eksik = False
        if not t["s400_str"] and not t["s600_str"]:
            veri_eksik = True
            detaylar.append("⚠️ DİKKAT VERİ EKSİK: Son400 ve Son600 bilgisi yok — sürpriz yapabilir!")

        # — ARAP ATLARI: mesafe farkı >200m → direkt elenir —
        # (en_iyi koşunun mesafesi bugünkü koşu mesafesinden >200m farklıysa)
        if is_arap:
            en_iyi_mes_str = en_iyi.get("son_mesafe", "")
            try:
                en_iyi_m = int(re.sub(r"[^0-9]", "", en_iyi_mes_str))
            except:
                en_iyi_m = 0
            if kosu_mesafe and en_iyi_m:
                fark_arap = abs(kosu_mesafe - en_iyi_m)
                if fark_arap > 200:
                    eleme_nedenleri.append(
                        f"Arap atı: en iyi koşu mesafesi çok farklı "
                        f"(bugün {kosu_mesafe}m, en iyi {en_iyi_m}m, Δ{fark_arap}m > 200m)"
                    )
                    detaylar.append(
                        f"❌ ARAP: En iyi koşu {en_iyi_m}m, bugün {kosu_mesafe}m (Δ{fark_arap}m) → DİREKT ELENDİ"
                    )
                else:
                    detaylar.append(f"✅ ARAP: Mesafe uygun ({en_iyi_m}m → {kosu_mesafe}m, Δ{fark_arap}m)")

        # — İNGİLİZ ATLARI: karma cinsiyette dişi at → direkt elenir —
        if is_ingiliz and karma_cinsiyet and horse.get("cinsiyet") == "dişi":
            eleme_nedenleri.append(
                f"Dişi İngiliz atı, erkeklerle karma koşuyor "
                f"({disi_sayisi} dişi, {erkek_sayisi} erkek aynı koşuda)"
            )
            detaylar.append("❌ İNGİLİZ DİŞİ: Erkeklerle karma koşu → DİREKT ELENDİ")

        s4 = t["s400_sn"]
        if s4 and ort_400:
            fark = ort_400 - s4   # pozitif = hızlı
            p4   = W_SON400 + round(fark * 15, 1)
            p4   = max(0, min(W_SON400 * 2, p4))
            puan += p4
            if fark >= 0.5:
                detaylar.append(f"✅ Son400 hızlı ({t['s400_str']}, ort:{sec_to_str(ort_400)})")
            elif fark <= -0.5:
                detaylar.append(f"❌ Son400 yavaş ({t['s400_str']}, ort:{sec_to_str(ort_400)})")
                eleme_nedenleri.append(f"Son400 yavaş ({t['s400_str']})")
            else:
                detaylar.append(f"➡️ Son400 ortalama ({t['s400_str']})")
        else:
            puan += W_SON400 // 2
            if not veri_eksik:
                detaylar.append("⚪ Son400 verisi yok")

        # — Son 600m —
        s6 = t["s600_sn"]
        if s6 and ort_600:
            fark = ort_600 - s6
            p6   = W_SON600 + round(fark * 10, 1)
            p6   = max(0, min(W_SON600 * 2, p6))
            puan += p6
            if fark >= 0.5:
                detaylar.append(f"✅ Son600 hızlı ({t['s600_str']})")
            elif fark <= -0.5:
                detaylar.append(f"❌ Son600 yavaş ({t['s600_str']})")
                eleme_nedenleri.append(f"Son600 yavaş ({t['s600_str']})")
            else:
                detaylar.append(f"➡️ Son600 ortalama ({t['s600_str']})")
        else:
            puan += W_SON600 // 2
            detaylar.append("⚪ Son600 verisi yok")

        # — Form trendi —
        genel_form = acc.get("genel_form","")
        if "YÜKSELİŞ" in genel_form:
            puan += W_FORM
            detaylar.append(f"✅ Form: {genel_form}")
        elif "DÜŞÜŞ" in genel_form:
            puan += 0
            detaylar.append(f"❌ Form: {genel_form}")
            eleme_nedenleri.append("Form düşüşte")
        elif genel_form:
            puan += W_FORM // 2
            detaylar.append(f"➡️ Form: {genel_form}")
        else:
            puan += W_FORM // 2
            detaylar.append("⚪ Form verisi yok")

        # — Mesafe uyumu (en_iyi koşuya göre) — tempo uyumsuzluğuyla birlikte eleme
        mesafe_uyumsuz = False
        en_iyi_mes_str = en_iyi.get("son_mesafe","")
        try:
            son_m = int(re.sub(r"[^0-9]","", en_iyi_mes_str))
        except:
            son_m = 0
        if kosu_mesafe and son_m:
            fark_m = abs(kosu_mesafe - son_m)
            if fark_m == 0:
                puan += W_MESAFE
                detaylar.append(f"✅ Aynı mesafe ({son_m}m)")
            elif fark_m <= 200:
                puan += round(W_MESAFE * 0.7)
                detaylar.append(f"➡️ Yakın mesafe ({son_m}m, Δ{fark_m}m)")
            elif fark_m <= 400:
                puan += round(W_MESAFE * 0.4)
                detaylar.append(f"⚠️ Farklı mesafe ({son_m}m, Δ{fark_m}m)")
                mesafe_uyumsuz = True
            else:
                puan += 0
                detaylar.append(f"❌ Çok farklı mesafe ({son_m}m, Δ{fark_m}m)")
                eleme_nedenleri.append(f"Mesafe uyumsuz ({son_m}m→{kosu_mesafe}m)")
                mesafe_uyumsuz = True
        else:
            puan += W_MESAFE // 2
            detaylar.append("⚪ Mesafe verisi yok")

        # — Koşu stili tespiti (TEMPO & SENARYO ile aynı mantık) —
        _rol = _at_roller.get(at_adi, {})
        _is_sprinter  = _rol.get("sprinter",  False)
        _is_geride_at = _rol.get("geride",    False)
        _is_on_grup   = _rol.get("on_grup",   False)
        _is_oyun_kur  = _rol.get("oyun_kurucu", False) or _rol.get("pot_oyun_kurucu", False)

        # Stil etiketi
        if _is_oyun_kur:
            stil_lbl = "OYN.KUR."
        elif _is_sprinter:
            stil_lbl = "SPRİNTER"
        elif _is_geride_at:
            stil_lbl = "GERİDE"
        elif _is_on_grup:
            stil_lbl = "ÖN GRUP"
        else:
            stil_lbl = "BELİRSİZ"

        # — STİL/TEMPO UYUM PUANI (%15) —
        # HIZLI tempo: sprinter veya geride karaktere +15 puan
        # YAVAS tempo: ön grup veya oyun kurucuya +15 puan
        if _tempo_karar == "HIZLI":
            if _is_sprinter or _is_geride_at:
                puan += W_STIL
                detaylar.append(f"✅ HIZLI tempoda avantajlı stil: {stil_lbl} (+{W_STIL}p)")
            elif _is_on_grup or _is_oyun_kur:
                puan += round(W_STIL * 0.4)
                detaylar.append(f"⚠️ HIZLI tempoda zor: {stil_lbl}")
            else:
                puan += W_STIL // 2
                detaylar.append(f"⚪ Stil belirsiz ({stil_lbl}), tarafsız puan")
        elif _tempo_karar == "YAVAS":
            if _is_on_grup or _is_oyun_kur:
                puan += W_STIL
                detaylar.append(f"✅ YAVAS tempoda avantajlı stil: {stil_lbl} (+{W_STIL}p)")
            elif _is_sprinter or _is_geride_at:
                puan += 0
                detaylar.append(f"❌ YAVAS tempoda dezavantajlı stil: {stil_lbl} (0p)")
            else:
                puan += W_STIL // 2
                detaylar.append(f"⚪ Stil belirsiz ({stil_lbl}), tarafsız puan")
        else:
            # Belirsiz tempo → tarafsız
            puan += W_STIL // 2
            detaylar.append(f"⚪ Tempo belirsiz, stil: {stil_lbl}")

        # — Madde 3: Yavaş tempoda GERİDE karakter + tempo uyumsuz → DİREKT ELENİ —
        yavas_geride_eleme = False
        if _tempo_karar == "YAVAS" and (_is_geride_at or _is_sprinter):
            tempo_bilgi_check = _tempo_uyum.get(at_adi, {"uyumlu": False})
            if not tempo_bilgi_check.get("uyumlu", False):
                yavas_geride_eleme = True
                eleme_nedenleri.append(
                    f"YAVAS tempo + GERİDE/SPRİNTER karakter ({stil_lbl}) + tempo uyumsuz → "
                    f"yavaş koşuda önde veya yakın olmayanlar kazanamaz → DİREKT ELENDİ"
                )

        # son 2 koşunun ortalama final sırası (ort_final_str için, puan artık STİL'den)
        son_obj    = acc.get("son", {})
        onceki_obj = acc.get("onceki", {})
        son_mes_key  = son_obj.get("son_mesafe","")
        onc_mes_key  = onceki_obj.get("son_mesafe","") or onceki_obj.get("mesafe","")
        son_final_sira = son_obj.get("sureler", {}).get(son_mes_key, {}).get("sira","")
        onc_final_sira = onceki_obj.get("sureler", {}).get(onc_mes_key, {}).get("sira","")
        gecerli_siralar = [x for x in [son_final_sira, onc_final_sira] if isinstance(x, int)]
        if gecerli_siralar:
            ort_final = sum(gecerli_siralar) / len(gecerli_siralar)
            ort_final_str = f"Ort.{ort_final:.1f} ({'/'.join(str(x) for x in gecerli_siralar)})"
        else:
            ort_final = None
            ort_final_str = "-"

        # — Sınıf avantajı bilgisi —
        sinif_bilgi = _sinif_avantaj.get(at_adi, {"var": False, "lbl": "-", "seviye": "bilinmiyor", "fark_hp": 0})
        if sinif_bilgi["var"]:
            detaylar.append(f"⭐ {sinif_bilgi['lbl']}")
        elif sinif_bilgi.get("seviye") == "dezavantaj":
            detaylar.append(f"⚠️ {sinif_bilgi['lbl']}")

        # — Tempo uyum etkisi —
        tempo_bilgi   = _tempo_uyum.get(at_adi, {"uyum_skoru":0,"uyumlu":False,"uyumsuz":False})
        tempo_uyumlu  = tempo_bilgi["uyumlu"]
        tempo_uyumsuz = tempo_bilgi["uyumsuz"]

        if tempo_uyumlu:
            puan += 20
            detaylar.append("✅ TEMPO UYUMLU+ (+20 bonus)")
        elif tempo_uyumsuz:
            detaylar.append("❌ TEMPO UYUMSUZ (her 2 koşu da uyumsuz)")

        # ── ELEME KARARI ──────────────────────────────────────────
        elendi = False

        # 1. KGS: 100+ gün start almamış → direkt elenir
        if kgs >= KGS_ESIK:
            elendi = True

        # 2. ARAP: en iyi koşu mesafesi >200m farklı → direkt elenir
        elif is_arap and any("Arap atı: en iyi koşu mesafesi çok farklı" in n for n in eleme_nedenleri):
            elendi = True

        # 3. İNGİLİZ DİŞİ: karma cinsiyet koşusunda → direkt elenir
        elif is_ingiliz and karma_cinsiyet and horse.get("cinsiyet") == "dişi":
            elendi = True

        # 4. Kilo: 1600m+ de 60kg üzeri VE 2+ kg artış → direkt elenir
        elif kosu_mesafe >= 1600 and at_kg > 60 and kg_artis >= 2:
            elendi = True

        # 5. YAVAS tempo + GERİDE karakter + tempo uyumsuz → direkt elenir
        elif yavas_geride_eleme:
            elendi = True

        else:
            son400_yavas = "Son400 yavaş" in " ".join(eleme_nedenleri)
            son600_yavas = "Son600 yavaş" in " ".join(eleme_nedenleri)
            her_ikisi_yavas = son400_yavas and son600_yavas

            # Tempo uyumsuz → direkt elenir (her zaman)
            if tempo_uyumsuz:
                elendi = True
                _tm = "1000m" if kosu_mesafe >= 1600 else "800m"
                acc_son = acc_data.get(at_adi, {}).get("son", {})
                tempo_sira_son = acc_son.get("sureler", {}).get(_tm, {}).get("sira", "")
                is_geride = isinstance(tempo_sira_son, int) and tempo_sira_son >= 5
                if is_geride:
                    eleme_nedenleri.append(
                        f"Tempo hızlı gidecek + GERİDE karakter ({_tm} sıra:{tempo_sira_son}) "
                        f"+ tempo uyumsuz (bu hızlı tempo için yetersiz kalır) → DİREKT ELENDİ"
                    )
                else:
                    eleme_nedenleri.append("Tempo uyumsuz (her 2 koşuda bu tempo çok hızlı) — DİREKT ELENDİ")

            # Mesafe uyumsuz VE tempo uyumu yok → ikisi birlikte direkt elenir
            elif mesafe_uyumsuz and not tempo_uyumlu:
                elendi = True
                eleme_nedenleri.append(
                    "Mesafe uyumsuz + tempo uyumu yok → DİREKT ELENDİ"
                )

            else:
                if her_ikisi_yavas and len(eleme_nedenleri) >= 3:
                    if tempo_uyumlu:
                        detaylar.append("⚠️ Son400+Son600 yavaş + başka kriter zayıf ama tempo uyumu korudu")
                    else:
                        elendi = True
                elif her_ikisi_yavas and len(eleme_nedenleri) >= 2:
                    if not tempo_uyumlu:
                        elendi = True
                elif len(eleme_nedenleri) >= 3:
                    if tempo_uyumlu:
                        detaylar.append(f"⚠️ {len(eleme_nedenleri)} kriter zayıf ama tempo uyumu korudu")
                    else:
                        elendi = True
                elif s4 and ort_400 and (s4 - ort_400) >= 1.0:
                    if not tempo_uyumlu:
                        elendi = True
                        eleme_nedenleri.append(f"Son400 çok yavaş (ort.+{s4-ort_400:.1f}sn)")

        # — Her iki koşuda tempo uyumlu → yıldız etiketi —
        tempo_yildiz = tempo_bilgi.get("uyum_skoru", 0) >= 3   # her 2 koşuda uyumlu + en az 1 başarı

        _s = {
            "no": horse["no"], "adi": at_adi,
            "puan": round(puan, 1),
            "detaylar": detaylar,
            "elendi": elendi,
            "eleme_nedenleri": eleme_nedenleri,
            "s400_str": t["s400_str"],
            "s600_str": t["s600_str"],
            "s400_sn":  sure_to_sec(t["s400_str"]) if t["s400_str"] else None,
            "s600_sn":  sure_to_sec(t["s600_str"]) if t["s600_str"] else None,
            "genel_form": genel_form,
            "ort_final_str": ort_final_str,
            "ort_final": ort_final,
            "tempo_uyumlu":  tempo_uyumlu,
            "tempo_uyumsuz": tempo_uyumsuz,
            "tempo_yildiz":  tempo_yildiz,
            "sinif_avantaj": sinif_bilgi["var"],
            "sinif_lbl":     sinif_bilgi["lbl"],
            "sinif_fark":    sinif_bilgi["fark_hp"],
            "kgs":           kgs,
            "at_kg":         at_kg,
            "veri_eksik":    veri_eksik,
            "stil_lbl":      stil_lbl,
            "jokey":         horse.get("jokey",""),
            "gp":            horse.get("gp",""),
            "hp":            horse.get("hp",""),
            "kg":            horse.get("kg",""),
            "agf":           horse.get("agf_pct",""),
            "cinsiyet":      horse.get("cinsiyet",""),
            "s400_en_iyi":   False,
            "s600_kotu":     False,
        }
        sonuclar.append(_s)

    # Puana göre sırala
    sonuclar.sort(key=lambda x: x["puan"], reverse=True)
    for i, s in enumerate(sonuclar, 1):
        s["sira"] = i

    # ── Madde 5: En kötü son600 olan 2 atı ele (elenmediyse) ─
    gec_s600 = [(i, sure_to_sec(s["s600_str"])) for i, s in enumerate(sonuclar)
                if s["s600_str"] and sure_to_sec(s["s600_str"]) is not None and not s["elendi"]]
    gec_s600_sirt = sorted(gec_s600, key=lambda x: x[1], reverse=True)
    _kotu_s600_idx = set(x[0] for x in gec_s600_sirt[:2])
    for idx in _kotu_s600_idx:
        if not sonuclar[idx]["elendi"]:
            sonuclar[idx]["elendi"] = True
            sonuclar[idx]["s600_kotu"] = True
            sonuclar[idx]["eleme_nedenleri"].append(
                f"Son600 en kötü 2 at içinde ({sonuclar[idx]['s600_str']}) → ELENDİ"
            )

    # ── Madde 6: En iyi son400 olan 3 atı işaretle ─
    gec_s400 = [(i, sure_to_sec(s["s400_str"])) for i, s in enumerate(sonuclar)
                if s["s400_str"] and sure_to_sec(s["s400_str"]) is not None]
    gec_s400_sirt = sorted(gec_s400, key=lambda x: x[1])
    _iyi_s400_idx = set(x[0] for x in gec_s400_sirt[:3])
    for idx in _iyi_s400_idx:
        sonuclar[idx]["s400_en_iyi"] = True

    # web_mode: ELEME sonuçlarını kaydet
    _web_sonuclar = sonuclar[:]

    # ── Sayfa başlığı ──────────────────────────────────────────
    ırk_lbl = ("ARAP" if is_arap else ("İNGİLİZ" if is_ingiliz else ""))
    cinsiyet_lbl = (f" | Karma Cinsiyet: {disi_sayisi}K {erkek_sayisi}E" if is_ingiliz and karma_cinsiyet else "")
    _senaryo_renk_map = {"HIZLI": "B71C1C", "YAVAS": "1B5E20", "BELIRSIZ": "546E7A"}
    _tempo_baslik_str = {"HIZLI": "⚡ BEKLENEN SENARYO: HIZLI", "YAVAS": "🐢 BEKLENEN SENARYO: YAVAS", "BELIRSIZ": "❓ SENARYO BELİRSİZ"}.get(_tempo_karar, "")
    eleme_title = (
        f"ELEME ANALİZİ  |  "
        f"{meta.get('sehir','').upper()} {meta.get('kosu_no','')}. KOŞU  |  "
        f"{meta.get('tarih_tr','')}  |  Mesafe: {kosu_mesafe}m"
        f"{(' | IRK: ' + ırk_lbl) if ırk_lbl else ''}{cinsiyet_lbl}"
        f"  |  {_tempo_baslik_str}"
    )
    tc = ws_e.cell(row=1, column=1, value=eleme_title)
    tc.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    tc.fill      = PatternFill("solid", start_color=_senaryo_renk_map.get(_tempo_karar, "880E4F"))
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_e.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws_e.row_dimensions[1].height = 24

    ac = ws_e.cell(row=2, column=1,
        value=(f"v9.8 | Puanlama: Son400 (%{W_SON400}) + Son600 (%{W_SON600}) + "
               f"Form (%{W_FORM}) + Mesafe Uyumu (%{W_MESAFE}) + Stil/Tempo Uyumu (%{W_STIL}) + "
               f"Tempo Uyumu (+20 bonus / uyumsuz=direkt elenir)  |  "
               f"EK ELEME: ARAP>200m mesafe | İngiliz dişi karma | KGS 100+ | 1600m+Kilo | Yavaş tempo+Geride  |  "
               f"🔴 Üzeri çizgili = ELENDİ  |  Puana göre sıralanmıştır"))
    ac.font      = Font(name="Arial", size=9, italic=True, color="555555")
    ac.alignment = Alignment(horizontal="left", vertical="center")
    ws_e.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws_e.row_dimensions[2].height = 14

    # ── Sütun başlıkları ───────────────────────────────────────
    _tempo_hdr_lbl = f"STİL\n({'HIZLI→SPR' if _tempo_karar=='HIZLI' else 'YAVAS→ÖNG' if _tempo_karar=='YAVAS' else 'BELİRSİZ'})"
    HDR_E = [
        ("SIRA",7), ("NO-AT ADI",22), ("PUAN",8),
        ("SON 400m",10), ("SON 600m",10), ("FORM",20),
        (_tempo_hdr_lbl, 14), ("DURUM",14), ("ELEME NEDENİ / DETAY",45)
    ]
    for ci, (lbl, w) in enumerate(HDR_E, 1):
        c = ws_e.cell(row=3, column=ci, value=lbl)
        c.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill      = PatternFill("solid", start_color="880E4F")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border("FFFFFF")
        ws_e.column_dimensions[get_column_letter(ci)].width = w
    ws_e.row_dimensions[3].height = 20

    # ── Veri satırları ─────────────────────────────────────────
    YUKSELIS_CLR = "1B5E20"
    DUSUS_CLR    = "B71C1C"
    STABIL_CLR   = "1565C0"

    for ri, s in enumerate(sonuclar):
        rn     = ri + 4
        even   = ri % 2 == 0
        elendi = s["elendi"]
        bg     = "FFEBEE" if elendi else ("EBF3FB" if even else WHITE)

        gform = s["genel_form"]
        if "YÜKSELİŞ" in gform:   form_clr = YUKSELIS_CLR
        elif "DÜŞÜŞ" in gform:     form_clr = DUSUS_CLR
        else:                       form_clr = STABIL_CLR

        # STİL rengi — tempo uyumuna göre
        _stil = s.get("stil_lbl", "-")
        if _tempo_karar == "HIZLI":
            if _stil in ("SPRİNTER", "GERİDE"):  stil_clr = YUKSELIS_CLR
            elif _stil in ("ÖN GRUP", "OYN.KUR."): stil_clr = DUSUS_CLR
            else: stil_clr = STABIL_CLR
        elif _tempo_karar == "YAVAS":
            if _stil in ("ÖN GRUP", "OYN.KUR."): stil_clr = YUKSELIS_CLR
            elif _stil in ("SPRİNTER", "GERİDE"): stil_clr = DUSUS_CLR
            else: stil_clr = STABIL_CLR
        else:
            stil_clr = STABIL_CLR

        if s["sira"] == 1:    puan_clr = "1B5E20"
        elif s["sira"] <= 3:  puan_clr = "2E7D32"
        elif elendi:          puan_clr = "B71C1C"
        else:                 puan_clr = "000000"

        # Yıldız: her 2 koşuda tempo uyumlu+başarılı
        yildiz = s.get("tempo_yildiz", False)
        at_label_e = f"⭐ {s['no']}-{s['adi']}" if yildiz else f"{s['no']}-{s['adi']}"
        at_label_clr = "1B5E20" if yildiz else "000000"

        durum_txt = "🔴 ELENDİ"   if elendi else (
            "⚠️ VERİ EKSİK" if s.get("veri_eksik") else (
            "🥇 ADAY" if s["sira"] <= 3 else "⚪ İZLEMEDE"))
        durum_clr = "B71C1C"      if elendi else (
            "E65100" if s.get("veri_eksik") else (
            "1B5E20" if s["sira"] <= 3 else "546E7A"))

        if elendi:
            detay_txt = " | ".join(s["eleme_nedenleri"])
        else:
            detay_txt = " | ".join(s["detaylar"][:3])

        satirlar = [
            (s["sira"],         False, "000000",     False),
            (at_label_e,        True,  at_label_clr, True),
            (s["puan"],         True,  puan_clr,     False),
            (s["s400_str"] or "-", False, "000000",  False),
            (s["s600_str"] or "-", False, "000000",  False),
            (gform or "-",      True,  form_clr,     False),
            (_stil or "-",      True,  stil_clr,     False),
            (durum_txt,         True,  durum_clr,    False),
            (detay_txt,         False, "555555",     True),
        ]

        for ci, (val, bold, clr, left) in enumerate(satirlar, 1):
            cx = ws_e.cell(row=rn, column=ci, value=val)
            # Madde 6: Son400 sütununu (ci==4) en iyi 3 ise kalın kırmızı yap
            if ci == 4 and s.get("s400_en_iyi"):
                cx.font = Font(name="Arial", size=10, bold=True, color="B71C1C",
                               strikethrough=elendi)
            else:
                cx.font = Font(name="Arial", size=10, bold=bold, color=clr,
                               strikethrough=elendi)
            cx.fill      = PatternFill("solid", start_color=bg)
            cx.alignment = Alignment(horizontal="left" if left else "center",
                                     vertical="center", wrap_text=(ci == 9))
            cx.border    = _border()

        if elendi:
            red = Side(style="thin", color="B71C1C")
            for ci in range(1, 10):
                ws_e.cell(row=rn, column=ci).border = Border(
                    left=red, right=red, top=red, bottom=red)

        ws_e.row_dimensions[rn].height = 18

    # ── Özet kutusu ───────────────────────────────────────────
    ozet_row = len(sonuclar) + 5
    ws_e.row_dimensions[ozet_row].height = 6   # boşluk

    adaylar   = [s for s in sonuclar if not s["elendi"]][:3]
    elenenler = [s for s in sonuclar if s["elendi"]]

    aday_str   = ", ".join(f"{s['no']}-{s['adi']} ({s['puan']}p)" for s in adaylar)
    elenen_str = ", ".join(f"{s['no']}-{s['adi']}" for s in elenenler) or "Yok"
    ozet_txt   = f"🏆 GÜÇLÜ ADAYLAR: {aday_str}   |   🔴 ELENENler: {elenen_str}"
    ozet_c = ws_e.cell(row=ozet_row + 1, column=1, value=ozet_txt)
    ozet_c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    ozet_c.fill      = PatternFill("solid", start_color="1B5E20")
    ozet_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ozet_c.border    = _border("1B5E20")
    ws_e.merge_cells(start_row=ozet_row + 1, start_column=1,
                     end_row=ozet_row + 1,   end_column=9)
    ws_e.row_dimensions[ozet_row + 1].height = 22

    # ══════════════════════════════════════════════════════════
    # ── TEMPO & SENARYO sayfası ────────────────────────────────
    # ══════════════════════════════════════════════════════════
    ws_t = wb.create_sheet("TEMPO & SENARYO")

    # Mesafeye göre tempo ölçüm noktası: 1600m+ → 1000m, kısa → 800m
    TEMPO_MES   = "1000m" if kosu_mesafe >= 1600 else "800m"
    TEMPO_TOL   = 1.5 if TEMPO_MES == "1000m" else 1.0
    BASARI_SIRA = 4

    # ── Her at için son 2 koşu verisini topla ─────────────────
    at_tempo = []
    for horse in horses:
        at_adi = horse["adi"]
        acc    = acc_data.get(at_adi, {})
        son    = acc.get("son", {})
        onceki = acc.get("onceki", {})

        def get_sira(veri, mes):
            return veri.get("sureler", {}).get(mes, {}).get("sira", "")

        def get_sure_sn(veri, mes):
            s = veri.get("sureler", {}).get(mes, {}).get("sure", "")
            return sure_to_sec(s) if s else None

        def get_final_sira(veri):
            mes = veri.get("son_mesafe","") or veri.get("mesafe","")
            return get_sira(veri, mes)

        # Sıra ve süre — TEMPO_MES'e göre (800m veya 1000m)
        s8_son   = get_sira(son,    TEMPO_MES)
        s8_onc   = get_sira(onceki, TEMPO_MES)
        sf_son   = get_final_sira(son)
        sf_onc   = get_final_sira(onceki)

        sure8_son = get_sure_sn(son,    TEMPO_MES)
        sure8_onc = get_sure_sn(onceki, TEMPO_MES)

        son400     = acc.get("son400","") or son.get("son400","")
        son600     = acc.get("son600","") or son.get("son600","")
        son400_sn  = sure_to_sec(son400) if son400 else None
        son600_sn  = sure_to_sec(son600) if son600 else None

        # En iyi (en kısa) süreyi bul — sıra yoksa bile süre varsa kullan
        best_sure = min(
            [s for s in [sure8_son, sure8_onc] if s is not None],
            default=None
        )

        # Kesin oyun kurucu: her iki koşuda da 1. sırada
        oyun_kurucu     = (s8_son == 1 and s8_onc == 1)
        # Potansiyel kurucu: en az birinde 1. sırada (başarısız olsa bile oyun kurucu sayılır)
        pot_oyun_kurucu = (s8_son == 1 or s8_onc == 1)
        # Süre bazlı aday: 1. sıra olmasa bile ilk 3 sıra içinde VE süresi var
        sure_bazli_aday = (
            not pot_oyun_kurucu and
            isinstance(s8_son, int) and s8_son <= 3 and
            sure8_son is not None
        )

        # ÖNE YAKIN: tempo mesafesinde sıra 1-4 (kurucu dahil)
        on_grup = (isinstance(s8_son, int) and s8_son <= 4)

        # UZUN SPRİNTER: tempo mesafesinde 5+ sırada (geride) + son koşusunda son400 hızlı
        # Son400 hızlı = genel ortalamayı burada bilmiyoruz, sonradan hesaplanacak
        # Şimdilik sadece "tempo mesafesinde geride" işaretleyelim, son400 karşılaştırması sonra
        geride = (isinstance(s8_son, int) and s8_son >= 5)

        at_tempo.append({
            "no": horse["no"], "adi": at_adi,
            "s8_son": s8_son,   "s8_onc": s8_onc,
            "sf_son": sf_son,   "sf_onc": sf_onc,
            "sure8_son": sure8_son, "sure8_onc": sure8_onc,
            "best_sure": best_sure,
            "son400": son400,   "son600": son600,
            "son400_sn": son400_sn, "son600_sn": son600_sn,
            "oyun_kurucu":     oyun_kurucu,
            "pot_oyun_kurucu": pot_oyun_kurucu,
            "sure_bazli_aday": sure_bazli_aday,
            "on_grup":         on_grup,
            "geride":          geride,
            "sprinter":        False,  # sonradan hesaplanacak
            "genel_form":      acc.get("genel_form",""),
        })

    # ── Tahmini tempo: sıra 1-3'teki atların her 2 koşusundaki ortalama süresi ─
    onde_sureler_t = []
    for t in at_tempo:
        for sure_, sira_ in [(t["sure8_son"], t["s8_son"]), (t["sure8_onc"], t["s8_onc"])]:
            if isinstance(sira_, int) and sira_ <= 3 and sure_ is not None:
                onde_sureler_t.append(sure_)
        # Süre bazlı aday da dahil
        if t["sure_bazli_aday"] and t["sure8_son"] is not None:
            onde_sureler_t.append(t["sure8_son"])

    # Aşırı uç değerleri temizle (en hızlı ve en yavaşın ortalaması)
    if onde_sureler_t:
        tahmini_tempo_sn  = sum(onde_sureler_t) / len(onde_sureler_t)
        tahmini_tempo_str = sec_to_str(tahmini_tempo_sn)
    else:
        tahmini_tempo_sn  = None
        tahmini_tempo_str = "?"

    # ── Son400 ortalaması → Sprinter tespiti için ─────────────
    # Uzun sprinter: tempo mesafesinde 5+ sırada + son400 ortalamanın altında (hızlı)
    gecerli_s400 = [t["son400_sn"] for t in at_tempo if t["son400_sn"]]
    ort_s400 = sum(gecerli_s400) / len(gecerli_s400) if gecerli_s400 else None

    for t in at_tempo:
        # Sprinter: tempo mesafesinde GERİDE (5+) VE son400 hızlı (ortalamadan kısa)
        s400_hizli = (t["son400_sn"] and ort_s400 and t["son400_sn"] < ort_s400)
        t["sprinter"] = t["geride"] and s400_hizli
        t["s400_hizli"] = s400_hizli

    # ── Her at için tempo uyum skoru hesapla ──────────────────
    # Mantık: atın geçmiş TEMPO_MES süresi <= tahmini_tempo + tolerans ise bu tempoyu kaldırabilir
    for t in at_tempo:
        uyum_skoru = 0
        uyum_detay = []

        for kosu_lbl, sure8, sf in [
            ("son koşu",    t["sure8_son"], t["sf_son"]),
            ("önceki koşu", t["sure8_onc"], t["sf_onc"]),
        ]:
            if sure8 is None or tahmini_tempo_sn is None:
                continue
            basarili = isinstance(sf, int) and sf <= BASARI_SIRA

            # At bu tempoyu KALDIRABİLİYOR MU?
            # sure8 <= tahmini_tempo + TEMPO_TOL → evet (daha hızlı veya yakın)
            # sure8 > tahmini_tempo + TEMPO_TOL → hayır (bu tempo çok hızlı kalır)
            fark_str  = f"{sure8 - tahmini_tempo_sn:+.1f}sn"  # negatif=hızlı, pozitif=yavaş
            kaldirabilir = sure8 <= tahmini_tempo_sn + TEMPO_TOL

            if kaldirabilir:
                hiz_notu = "daha hızlı" if sure8 < tahmini_tempo_sn else "yakın"
                if basarili:
                    uyum_skoru += 2
                    uyum_detay.append(
                        f"{kosu_lbl}: {TEMPO_MES} {sec_to_str(sure8)} ({fark_str} tempo, {hiz_notu}) → "
                        f"final {sf}. UYUMLU+BASARILI")
                else:
                    uyum_skoru += 1
                    uyum_detay.append(
                        f"{kosu_lbl}: {TEMPO_MES} {sec_to_str(sure8)} ({fark_str} tempo, {hiz_notu}) → "
                        f"final {sf}. uyumlu ama finale giremedi")
            else:
                uyum_detay.append(
                    f"{kosu_lbl}: {TEMPO_MES} {sec_to_str(sure8)} ({fark_str} tempo) → "
                    f"bu tempo çok hızlı, yetersiz kalabilir")

        t["uyum_skoru"] = uyum_skoru
        t["uyum_detay"] = uyum_detay
        t["tempo_uyumlu"] = uyum_skoru >= 2

    # ── Gruplar ───────────────────────────────────────────────
    oyun_kurucular  = [t for t in at_tempo if t["oyun_kurucu"]]
    pot_kurucular   = [t for t in at_tempo if t["pot_oyun_kurucu"] and not t["oyun_kurucu"]]
    # Süre bazlı adaylar: sıra 1 olmasa bile en hızlı süreli öndeki atlar
    sure_adaylar    = [t for t in at_tempo if t["sure_bazli_aday"]]
    # En hızlı süreye göre sırala
    sure_adaylar    = sorted(sure_adaylar, key=lambda x: x["sure8_son"] if x["sure8_son"] else 999)
    sprinterlar     = [t for t in at_tempo if t["sprinter"]]
    on_gruplar      = [t for t in at_tempo if t["on_grup"]]

    # Eğer hiç 1. sıra alan yoksa, sure_adaylardan en hızlısını pot. kurucu say
    if not oyun_kurucular and not pot_kurucular and sure_adaylar:
        # En hızlı sure_aday'ı pot. kurucu olarak işaretle
        for t in at_tempo:
            if t["adi"] == sure_adaylar[0]["adi"]:
                t["pot_oyun_kurucu"] = True
                t["sure_bazli"] = True   # Süre bazlı seçildiğini işaretle
        pot_kurucular = [sure_adaylar[0]]

    # ── Tempo kararı ──────────────────────────────────────────
    # Oyun kurucu: son 2 koşuda da 1. sırada
    # Pot. kurucu: en az birinde 1. sırada VEYA en hızlı süreye sahip öndeki at

    if len(oyun_kurucular) >= 2:
        tempo_karar = "HIZLI"
        tempo_renk  = "B71C1C"
        kurucu_adlar = ", ".join(f"{t['no']}-{t['adi']}" for t in oyun_kurucular)
        tempo_detay = (f"{kurucu_adlar} birbirini sikistiracak → tempo yukselir, yipranma olur. "
                       f"Tahm. {TEMPO_MES}: {tahmini_tempo_str}. "
                       f"Son anlarda patlayan (Son400 hizli + geride) atlar avantajli.")
        rol_sanslilar    = [t for t in sprinterlar if t["tempo_uyumlu"]]
        rol_dezavantajli = [t for t in on_gruplar  if not t["tempo_uyumlu"] and not t["oyun_kurucu"]]

    elif len(oyun_kurucular) == 1:
        ok = oyun_kurucular[0]
        rakipler = [t for t in pot_kurucular
                    if isinstance(t["s8_son"], int) and t["s8_son"] <= 3]
        if rakipler:
            rakip_adlar = ", ".join(f"{t['no']}-{t['adi']}" for t in rakipler)
            tempo_karar = "HIZLI"
            tempo_renk  = "E65100"
            tempo_detay = (f"{ok['no']}-{ok['adi']} one gider, {rakip_adlar} baski yapar → tempo yukselir. "
                           f"Tahm. {TEMPO_MES}: {tahmini_tempo_str}. "
                           f"Son anlarda patlayan (Son400 hizli + geride) atlar avantajli.")
            rol_sanslilar    = [t for t in sprinterlar if t["tempo_uyumlu"]]
            rol_dezavantajli = [t for t in on_gruplar  if not t["tempo_uyumlu"] and not t["oyun_kurucu"]]
        else:
            tempo_karar = "YAVAS"
            tempo_renk  = "1B5E20"
            tempo_detay = (f"{ok['no']}-{ok['adi']} tek basina onde gidecek, rakipsiz. "
                           f"Tahm. {TEMPO_MES}: {tahmini_tempo_str}. "
                           f"Tempo kontrollü → onde veya ona yakin olan atlar avantajli.")
            rol_sanslilar    = [t for t in at_tempo
                                if (t["oyun_kurucu"] or t["on_grup"]) and t["tempo_uyumlu"]]
            rol_dezavantajli = [t for t in sprinterlar if not t["tempo_uyumlu"]]

    elif len(pot_kurucular) == 1:
        ok = pot_kurucular[0]
        sure_notu = f" (en hizli {TEMPO_MES} suresi: {sec_to_str(ok['sure8_son'])})" if ok.get("sure_bazli") else ""
        tempo_karar = "YAVAS"
        tempo_renk  = "1B5E20"
        tempo_detay = (f"{ok['no']}-{ok['adi']} muhtemel kurucu{sure_notu}, rakipsiz gidebilir. "
                       f"Tahm. {TEMPO_MES}: {tahmini_tempo_str}. "
                       f"Tempo kontrollü → onde veya ona yakin olan atlar avantajli.")
        rol_sanslilar    = [t for t in at_tempo
                            if (t["pot_oyun_kurucu"] or t["on_grup"]) and t["tempo_uyumlu"]]
        rol_dezavantajli = [t for t in sprinterlar if not t["tempo_uyumlu"]]

    elif len(pot_kurucular) >= 2:
        pot_sirali   = sorted(pot_kurucular,
                              key=lambda x: x["sure8_son"] if x["sure8_son"] else 999)
        hizli_kurucu = pot_sirali[0]
        kurucu_adlar = ", ".join(f"{str(t['no'])}-{t['adi']}" for t in pot_sirali[:3])
        tempo_karar  = "HIZLI"
        tempo_renk   = "E65100"
        tempo_detay  = (f"Birden fazla muhtemel kurucu: {kurucu_adlar}. "
                        f"En hizlisi {hizli_kurucu['no']}-{hizli_kurucu['adi']} tempoyu belirler, "
                        f"digerleri sikistirir → yipranma olursa son anlarda patlayan avantajli. "
                        f"Tahm. {TEMPO_MES}: {tahmini_tempo_str}.")
        rol_sanslilar    = [t for t in sprinterlar if t["tempo_uyumlu"]]
        rol_dezavantajli = []

    else:
        tempo_karar = "BELIRSIZ"
        tempo_renk  = "546E7A"
        tempo_detay = "Onde gidecek at belirlenemiyor — veri yetersiz."
        rol_sanslilar    = [t for t in at_tempo if t["tempo_uyumlu"]]
        rol_dezavantajli = []

    sanslı_adlar = [f"{t['no']}-{t['adi']}" for t in rol_sanslilar]
    dezav_adlar  = [f"{t['no']}-{t['adi']}" for t in rol_dezavantajli]

    # ── web_mode: tempo verisini kaydet ───────────────────────
    _web_tempo = {
        "karar":           tempo_karar,
        "aciklama":        tempo_detay,
        "tahmini":         tahmini_tempo_str,
        "tempo_mes":       TEMPO_MES,
        "sanslilar":       sanslı_adlar,
        "dezavantajlilar": dezav_adlar,
        "oyun_kurucular":  [f"{t['no']}-{t['adi']}" for t in oyun_kurucular],
        "pot_kurucular":   [f"{t['no']}-{t['adi']}" for t in pot_kurucular],
        "sprinterlar":     [f"{t['no']}-{t['adi']}" for t in sprinterlar],
        "on_gruplar":      [f"{t['no']}-{t['adi']}" for t in on_gruplar],
        "geride_adlar":    [f"{t['no']}-{t['adi']}" for t in at_tempo
                            if t["geride"] and not t["sprinter"]],
        "at_detay":        [
            {
                "no": t["no"], "adi": t["adi"],
                "rol": ("OYN.KUR." if t["oyun_kurucu"]
                        else "POT.KUR." if t["pot_oyun_kurucu"]
                        else "SPRİNTER" if t["sprinter"]
                        else "ÖN GRUP" if t["on_grup"]
                        else "GERİDE"  if t["geride"]
                        else "BELİRSİZ"),
                "uyumlu":    t.get("tempo_uyumlu", False),
                "uyum_detay": t.get("uyum_detay", []),
                "son400":    t.get("son400",""),
                "son600":    t.get("son600",""),
            }
            for t in at_tempo
        ],
    }

    # web_mode: Excel yazmadan veri döndür
    if web_mode:
        return {
            "eleme":  _web_sonuclar,
            "tempo":  _web_tempo,
        }
    tempo_title = (
        f"TEMPO & SENARYO ANALIZI  |  "
        f"{meta.get('sehir','').upper()} {meta.get('kosu_no','')}. KOSU  |  "
        f"{meta.get('tarih_tr','')}  |  Mesafe: {kosu_mesafe}m"
    )
    tc = ws_t.cell(row=1, column=1, value=tempo_title)
    tc.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    tc.fill      = PatternFill("solid", start_color=tempo_renk)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_t.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws_t.row_dimensions[1].height = 26

    tk = ws_t.cell(row=2, column=1,
        value=f"MUHTEMEL TEMPO: {tempo_karar}  |  Tahmini {TEMPO_MES}: {tahmini_tempo_str}  |  {tempo_detay}")
    tk.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    tk.fill      = PatternFill("solid", start_color=tempo_renk)
    tk.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_t.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws_t.row_dimensions[2].height = 30

    sans_txt = (
        f"SANSLI (bu tempoda basarili): "
        f"{', '.join(sanslı_adlar) if sanslı_adlar else 'Belirlenemedi'}   |   "
        f"RISKLI (bu tempoya uyumsuz): "
        f"{', '.join(dezav_adlar) if dezav_adlar else 'Yok'}"
    )
    sk = ws_t.cell(row=3, column=1, value=sans_txt)
    sk.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    sk.fill      = PatternFill("solid", start_color="1A237E")
    sk.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_t.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
    ws_t.row_dimensions[3].height = 22

    ac_txt = (
        f"Oyun Kurucu: Son 2 kosuda {TEMPO_MES} 1. sira  |  "
        f"Sprinter: {TEMPO_MES} 5+ ama final ilk 4  |  "
        f"On Grup: {TEMPO_MES} 2-4. sira  |  "
        f"Tempo Uyumu: Gecmis {TEMPO_MES} suresi tahmini tempoya {TEMPO_TOL}sn toleransla uyuyor + final ilk {BASARI_SIRA}"
    )
    ac = ws_t.cell(row=4, column=1, value=ac_txt)
    ac.font      = Font(name="Arial", size=9, italic=True, color="333333")
    ac.fill      = PatternFill("solid", start_color="F3E5F5")
    ac.alignment = Alignment(horizontal="left", vertical="center")
    ws_t.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11)
    ws_t.row_dimensions[4].height = 14

    # ── Sütun başlıkları ───────────────────────────────────────
    HDR_T = [
        ("NO-AT ADI",22),
        ("ROL",15),
        ("800m SIRA\n(SON)",9),
        ("800m SIRA\n(ONCEKI)",9),
        ("FINAL\n(SON)",8),
        ("FINAL\n(ONCEKI)",8),
        ("SON 400m",9),
        ("SON 600m",9),
        ("TEMPO\nUYUMU",10),
        ("FORM",16),
        ("SENARYO & TEMPO UYUM DETAYI",55),
    ]
    for ci, (lbl, w) in enumerate(HDR_T, 1):
        c = ws_t.cell(row=5, column=ci, value=lbl)
        c.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill      = PatternFill("solid", start_color="4A148C")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border("FFFFFF")
        ws_t.column_dimensions[get_column_letter(ci)].width = w
    ws_t.row_dimensions[5].height = 30

    # ── Veri satırları ─────────────────────────────────────────
    def rol_sirasi(t):
        if t["oyun_kurucu"]:       return (0, -t["uyum_skoru"])
        if t["pot_oyun_kurucu"]:   return (1, -t["uyum_skoru"])
        if t["on_grup"]:           return (2, -t["uyum_skoru"])
        if t["sprinter"]:          return (3, -t["uyum_skoru"])
        return (4, -t["uyum_skoru"])

    at_sirali = sorted(at_tempo, key=rol_sirasi)

    for ri, t in enumerate(at_sirali):
        rn = ri + 6

        if t["oyun_kurucu"]:
            rol = "OYUN KURUCU"; rol_bg = "FFEBEE"; rol_clr = "B71C1C"
        elif t.get("sure_bazli"):
            rol = "SURE BAZLI KURUCU"; rol_bg = "FFF8E1"; rol_clr = "F57F17"
        elif t["pot_oyun_kurucu"]:
            rol = "POT. KURUCU"; rol_bg = "FFF3E0"; rol_clr = "E65100"
        elif t["sprinter"]:
            rol = "SPRINTER";    rol_bg = "E8F5E9"; rol_clr = "1B5E20"
        elif t["on_grup"]:
            rol = "ON GRUP";     rol_bg = "E3F2FD"; rol_clr = "1565C0"
        else:
            rol = "GERIDE";      rol_bg = "F5F5F5"; rol_clr = "546E7A"

        # Tempo uyum etiketi
        if t["uyum_skoru"] >= 2:
            uyum_lbl = "UYUMLU+"; uyum_clr = "1B5E20"
        elif t["uyum_skoru"] == 1:
            uyum_lbl = "KISMI";   uyum_clr = "E65100"
        else:
            uyum_lbl = "UYUMSUZ"; uyum_clr = "B71C1C"

        # Senaryo
        uyum_acik = " | ".join(t["uyum_detay"]) if t["uyum_detay"] else "Veri yok"
        if t["oyun_kurucu"] and len(oyun_kurucular) >= 2:
            rol_acik = (f"Son 2 kosuda {TEMPO_MES}'de 1. sira → kesin kurucu. "
                        f"Diger kurucularla sikisacak → yipranma riski yuksek. "
                        f"Tempo yuksek giderse son anlarda patlayan atlar devreye girer. ")
        elif t["oyun_kurucu"]:
            rol_acik = (f"Son 2 kosuda {TEMPO_MES}'de 1. sira → tek kurucu, rakipsiz. "
                        f"Tempo kontrollü gidecek → onde veya ona yakin olan kazanabilir. ")
        elif t.get("sure_bazli"):
            rol_acik = (f"{TEMPO_MES}'de {t['s8_son']}. sirada ama en hizli sure ({sec_to_str(t['sure8_son'])}) → "
                        f"muhtemelen one gecip tempoyu belirleyecek. "
                        f"Rakipsiz giderse onde kalan kazanir. ")
        elif t["pot_oyun_kurucu"]:
            rol_acik = (f"Son kosulardan birinde {TEMPO_MES}'de 1. siradaydi → one cikabilir. "
                        f"Karsisinda rakip varsa tempo yukselir, yoksa kontrollü gider. ")
        elif t["sprinter"]:
            s8 = t['s8_son'] if isinstance(t['s8_son'], int) else '?'
            rol_acik = (f"{TEMPO_MES}'de {s8}. sirada (geride) + Son400 hizli → son anlarda patlayan. "
                        f"TEMPO YUKSEK GIDERSE SANSLI — onde catisma olursa son anlarda fırlar. ")
        elif t["on_grup"]:
            s8 = t['s8_son'] if isinstance(t['s8_son'], int) else '?'
            rol_acik = (f"{TEMPO_MES}'de {s8}. sirada → onde yakin. "
                        f"TEMPO YAVAS/KONTROLLU GIDERSE SANSLI — onde yalniz kalacak at yakinindaysa tehlikeli. ")
        else:
            s8 = t['s8_son'] if isinstance(t['s8_son'], int) else '?'
            sf = t['sf_son'] if isinstance(t['sf_son'], int) else '?'
            rol_acik = (f"{TEMPO_MES}'de {s8}. sirada, final'de {sf}. sira. "
                        f"Belirgin avantaj yok. ")

        senaryo = rol_acik + " TEMPO UYUMU: " + uyum_acik

        gf = t["genel_form"]
        if "YÜKSELİŞ" in gf:   fc = "1B5E20"
        elif "DÜŞÜŞ" in gf:     fc = "B71C1C"
        else:                    fc = "1565C0"

        def ss(v): return str(v) if isinstance(v, int) else "-"
        s8s_clr = "B71C1C" if t["s8_son"]==1 else ("E65100" if isinstance(t["s8_son"],int) and t["s8_son"]<=3 else "000000")
        s8o_clr = "B71C1C" if t["s8_onc"]==1 else ("E65100" if isinstance(t["s8_onc"],int) and t["s8_onc"]<=3 else "000000")
        sfs_clr = "1B5E20" if isinstance(t["sf_son"],int) and t["sf_son"]<=BASARI_SIRA else "000000"
        sfo_clr = "1B5E20" if isinstance(t["sf_onc"],int) and t["sf_onc"]<=BASARI_SIRA else "000000"

        cols = [
            (f"{t['no']}-{t['adi']}", True,  "000000",  True),
            (rol,                      True,  rol_clr,   False),
            (ss(t["s8_son"]),          True,  s8s_clr,   False),
            (ss(t["s8_onc"]),          True,  s8o_clr,   False),
            (ss(t["sf_son"]),          True,  sfs_clr,   False),
            (ss(t["sf_onc"]),          True,  sfo_clr,   False),
            (t["son400"] or "-",       False, "000000",  False),
            (t["son600"] or "-",       False, "000000",  False),
            (uyum_lbl,                 True,  uyum_clr,  False),
            (gf or "-",                True,  fc,        False),
            (senaryo,                  False, "333333",  True),
        ]
        for ci, (val, bold, clr, left) in enumerate(cols, 1):
            cx = ws_t.cell(row=rn, column=ci, value=val)
            cx.font      = Font(name="Arial", size=10, bold=bold, color=clr)
            cx.fill      = PatternFill("solid", start_color=rol_bg)
            cx.alignment = Alignment(horizontal="left" if left else "center",
                                     vertical="center", wrap_text=(ci==11))
            cx.border    = _border()
        ws_t.row_dimensions[rn].height = 40


    if not web_mode:
        wb.save(fname)




# ─── ANA AKIŞ ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  LİDERFORM + ACCURACE TEK KOŞU EXCEL OLUŞTURUCU  v9.7")
    print("  - basari_skoru() ile akıllı en iyi koşu seçimi")
    print("  - Cinsiyet tespiti (kısrak/dişi/aygır/erkek)")
    print("  - Arap atları: mesafe farkı >200m ise direkt elenir")
    print("  - İngiliz atları: karma cinsiyette dişi at elenir")
    print("=" * 60)

    if len(sys.argv) > 1:
        url = sys.argv[1].strip()
    else:
        url = input(
            "\nLütfen koşu linkini girin\n"
            "(Örnek: https://liderform.com.tr/program/2026-03-28/istanbul/1)\n> "
        ).strip()

    if "liderform.com.tr" not in url:
        print("HATA: Geçerli bir liderform.com.tr linki girin.")
        sys.exit(1)

    meta = url_meta(url)
    if not meta:
        print("HATA: Link formatı hatalı.")
        sys.exit(1)

    print(f"\n  Tarih  : {meta['tarih_tr']}")
    print(f"  Şehir  : {meta['sehir']}")
    print(f"  Koşu   : {meta['kosu_no']}. Koşu")

    # ── 1. Liderform program sayfası ──
    print(f"\n[1/3] Liderform program sayfası çekiliyor...")
    try:
        prog_url = (f"https://liderform.com.tr/program/"
                    f"{meta['tarih_iso']}/{meta['sehir_raw']}/{meta['kosu_no']}")
        soup   = fetch_soup(prog_url)
        rinfo  = race_info(soup)
        horses = parse_ana_sayfa(soup)
    except Exception as e:
        print(f"HATA: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)

    if not horses:
        print("UYARI: At verisi bulunamadı!")
        sys.exit(1)

    print(f"       ✓ {len(horses)} at | "
          f"{rinfo['tip']} | {rinfo['pist']} | {rinfo['mesafe']}")

    # ── 2. Performans sayfasından her atın gerçek son koşu linkini al ──
    print(f"\n[2/3] Performans sayfasından son koşu bilgileri çekiliyor...")
    son_kosu_map = son_kosu_linklerini_cek(
        meta["tarih_iso"], meta["sehir_raw"], meta["kosu_no"]
    )
    print(f"       ✓ {len(son_kosu_map)} at için son koşu bulundu")

    # ── 3. Her at için AYRI Accurace verisi çek ──
    print(f"\n[3/4] Her atın son koşusu için Accurace verisi çekiliyor...")
    acc_data = fetch_accurace_per_horse(horses, son_kosu_map)

    basarili = sum(1 for v in acc_data.values() if v.get("en_iyi", {}).get("sureler"))
    print(f"       ✓ {basarili}/{len(horses)} at için Accurace verisi alındı")

    # ── 3. Excel ──
    safe  = re.sub(r"[^\w]", "_", meta["sehir"])
    fname = f"{meta['tarih_iso']}_{safe}_{meta['kosu_no']}_kosu_v9_7.xlsx"

    print(f"\n[4/4] Excel oluşturuluyor: {fname}")
    try:
        build_excel(meta, rinfo, horses, acc_data, [], fname, son_kosu_map)
        print(f"       Kaydedildi → {os.path.abspath(fname)}")
    except Exception as e:
        print(f"\n❌ Excel oluşturma HATASI: {e}")
        import traceback; traceback.print_exc()
        # Yine de VERİ ve ANALİZ sayfalarını kaydet
        print("\n⚠ Sadece VERİ ve ANALİZ sayfaları kaydediliyor...")
        from openpyxl import Workbook as WB2
        wb2 = WB2()
        ws2 = wb2.active; ws2.title = "VERİ"
        write_block(ws2, meta, rinfo, horses, start_row=2, start_col=1)
        ws2b = wb2.create_sheet("ANALİZ")
        write_block(ws2b, meta, rinfo, horses, start_row=6, start_col=6)
        fname2 = fname.replace(".xlsx","_temel.xlsx")
        wb2.save(fname2)
        print(f"       Kaydedildi → {os.path.abspath(fname2)}")

    # Terminal özet
    print()
    print(f"{'NO':>3}  {'AT ADI':<16}  {'GP':>4}  {'HP':>4}  {'KG':>4}  "
          f"{'JOKEY':<12}  {'YRC':>3}  {'ST':>3}  {'KGS':>4}  {'AGF %':>8}")
    print("─" * 80)
    for h in horses:
        print(f"{str(h['no']):>3}  {str(h['adi']):<16}  {str(h['gp']):>4}  "
              f"{str(h['hp']):>4}  {str(h['kg']):>4}  {str(h['jokey']):<12}  "
              f"{str(h['yrc']):>3}  {str(h['st']):>3}  {str(h['kgs']):>4}  "
              f"{str(h['agf_pct']):>8}")
    print("─" * 80)

    print("\nACCURACE EN İYİ KOŞU ÖZETİ (basari_skoru ile seçildi):")
    print(f"{'AT':<18}  {'ETİKET':>8}  {'FİNAL SÜRE':>12}  {'FİNAL SIRA':>10}")
    print("─" * 60)
    for h in horses:
        acc        = acc_data.get(h["adi"], {})
        en_iyi     = acc.get("en_iyi", {})
        etik       = acc.get("en_iyi_etik", "-")
        son_mesafe = en_iyi.get("son_mesafe", "-")
        sureler    = en_iyi.get("sureler", {})
        final      = sureler.get(son_mesafe, {})
        sure       = final.get("sure", "-")
        sira       = final.get("sira", "-")
        print(f"{h['adi']:<18}  {etik:>8}  {str(sure):>12}  {str(sira):>10}")

    print(f"\n✓ Tamamlandı! → {fname}\n")
    print("Excel sayfaları:")
    print("  1. VERİ          → Program verileri (A2'den)")
    print("  2. ANALİZ        → Program verileri (F6'dan)")
    print("  3. ACCURACE      → Her atın son+önceki koşusu hızlanma verisi")
    print("  4. ELEME ANALİZİ → Puanlama ve eleme (ırk/cinsiyet+tempo bazlı)")
    print("  5. TEMPO & SENARYO → Oyun kurucu ve tempo analizi")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProgram durduruldu.")
    except Exception as e:
        print(f"\nHata: {e}")
        import traceback; traceback.print_exc()
